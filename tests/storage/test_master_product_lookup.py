from __future__ import annotations

from decimal import Decimal
import json
from pathlib import Path
import shutil
import sqlite3
import tempfile
import unittest
from unittest.mock import patch
from uuid import uuid4

import numpy as np
import pandas as pd
from pandas.testing import assert_frame_equal
from openpyxl import Workbook, load_workbook

from data_pipeline.contracts import normalize_integer_code
from data_pipeline.ingest import SOURCE_BATCH_COLUMNS
from data_pipeline.storage import master_product_lookup as lookup_module
from data_pipeline.storage.master_product_lookup import (
    DATABASE_FILENAME,
    EmptyMasterLookupError,
    MANIFEST_FILENAME,
    MasterLookupConflictError,
    MasterLookupIntegrityError,
    MasterLookupStorageError,
    MasterSheetDiscoveryError,
    MasterSheetSchemaError,
    SupplyBatchJoinError,
    build_master_product_lookup,
    create_master_lineage,
    discover_master_sheets,
    join_supply_batch_to_master,
    open_master_product_lookup,
    stream_master_product_keys,
    verify_master_product_lookup,
)


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
TEMP_PARENT = Path(tempfile.gettempdir())
MASTER_HEADERS = ["의료기기품목일련번호", "모델일련번호", "UDIDI일련번호", "ignored"]
MOJIBAKE_HEADERS = [
    "?\uc10e\uc9ba\u6e72\uaccc\ub9b0?\ub348\u3049?\uc1f0\uc834\u8e30\ub34a\uc0c7",
    "\uf9cf\u2464\ub73d?\uc1f0\uc834\u8e30\ub34a\uc0c7",
    "UDIDI?\uc1f0\uc834\u8e30\ub34a\uc0c7",
]


def write_master(
    path: Path,
    sheets: list[tuple[str, list[list[object]]]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets:
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)
    workbook.close()


def default_master(path: Path, *, duplicate: bool = False) -> Path:
    sheets: list[tuple[str, list[list[object]]]] = [
        ("overview", [["synthetic metadata"], ["not a product-key header"]])
    ]
    for number, name in enumerate(("z-sheet", "a-sheet", "m-sheet"), start=1):
        rows: list[list[object]] = [["synthetic metadata"], MASTER_HEADERS]
        rows.append([number, number * 10, number * 100, "not stored"])
        if duplicate and number == 1:
            rows.append([number, number * 10, number * 100, "duplicate"])
        sheets.append((name, rows))
    write_master(path, sheets)
    return path


def supply_batch(keys: list[tuple[object, object, object]]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for position, (item, model, udi) in enumerate(keys):
        rows.append(
            {
                "supply_date": pd.Timestamp("2026-01-15"),
                "src_company_id": "co:10",
                "dst_company_id": "co:20",
                "item_serial": item,
                "model_serial": model,
                "udi_serial": udi,
                "item_group_id": "SYNTHETIC-GROUP",
                "item_name_id": "SYNTHETIC-ITEM",
                "transaction_type": "SUPPLY",
                "amount_clean": Decimal("1"),
                "raw_supply_qty": Decimal("1"),
                "piece_qty": Decimal("1"),
                "udi": f"SYNTHETIC-{position}",
                "supplier_type": "SYNTHETIC",
                "receiver_type": "SYNTHETIC",
                "supplier_region": "00",
                "receiver_region": "00",
                "source_version": "nids-supply-v1:" + "a" * 64,
                "source_row_id": f"nids-row-v1:{position:064x}",
                "row_quality_flags": "",
            }
        )
    return pd.DataFrame(rows, columns=SOURCE_BATCH_COLUMNS)


class TrackingWorkbook:
    def __init__(self, workbook: object) -> None:
        self.workbook = workbook
        self.close_calls = 0

    @property
    def sheetnames(self) -> list[str]:
        return self.workbook.sheetnames

    def __getitem__(self, name: str) -> object:
        return self.workbook[name]

    def close(self) -> None:
        self.close_calls += 1
        self.workbook.close()


class MasterProductLookupTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TEMP_PARENT / f".master-lookup-{uuid4().hex[:10]}"
        self.temp_dir.mkdir(parents=True)
        self.lookup_root = self.temp_dir / "lookup"
        self.master_path = default_master(self.temp_dir / "synthetic-master.xlsx")

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def build(self):
        return build_master_product_lookup([self.master_path], self.lookup_root)

    def final_dir(self, source_hash: str) -> Path:
        return (
            self.lookup_root
            / "master_product_lookup"
            / "schema_version=1.0.0"
            / f"source_hash={source_hash}"
        )

    def test_discovers_all_three_sheets_with_shifted_headers_in_sorted_order(self) -> None:
        discovered = discover_master_sheets(self.master_path)
        self.assertEqual([sheet.name for sheet in discovered], ["a-sheet", "m-sheet", "z-sheet"])
        self.assertTrue(all(sheet.header_row == 2 for sheet in discovered))

    def test_only_confirmed_utf8_master_headers_are_supported(self) -> None:
        self.assertEqual(
            lookup_module.MASTER_HEADER_ALIASES,
            {
                "item_serial": ("의료기기품목일련번호",),
                "model_serial": ("모델일련번호",),
                "udi_serial": ("UDIDI일련번호",),
            },
        )
        self.assertEqual(len(discover_master_sheets(self.master_path)), 3)

    def test_mojibake_headers_are_not_treated_as_supported_aliases(self) -> None:
        path = self.temp_dir / "mojibake.xlsx"
        write_master(path, [("data", [MOJIBAKE_HEADERS, [1, 2, 3]])])
        with self.assertRaises(MasterSheetDiscoveryError):
            discover_master_sheets(path)

    def test_missing_key_header_is_rejected_before_data_rows(self) -> None:
        path = self.temp_dir / "missing.xlsx"
        write_master(path, [("data", [["intro"], MASTER_HEADERS[:2], [1, 2]])])
        with self.assertRaisesRegex(MasterSheetSchemaError, "udi_serial"):
            discover_master_sheets(path)

    def test_duplicate_and_ambiguous_headers_are_rejected(self) -> None:
        duplicate = self.temp_dir / "duplicate-header.xlsx"
        headers = [MASTER_HEADERS[0], MASTER_HEADERS[1], MASTER_HEADERS[2], MASTER_HEADERS[2]]
        write_master(duplicate, [("data", [headers, [1, 2, 3, 3]])])
        with self.assertRaises(MasterSheetDiscoveryError):
            discover_master_sheets(duplicate)
        ambiguous = self.temp_dir / "ambiguous.xlsx"
        write_master(ambiguous, [("data", [MASTER_HEADERS, MASTER_HEADERS, [1, 2, 3]])])
        with self.assertRaisesRegex(MasterSheetDiscoveryError, "Ambiguous"):
            discover_master_sheets(ambiguous)

    def test_lineage_is_path_and_input_order_independent_and_master_namespaced(self) -> None:
        first = self.temp_dir / "a.xlsx"
        second = self.temp_dir / "b.xlsx"
        default_master(first)
        default_master(second)
        moved = self.temp_dir / "nested"
        moved.mkdir()
        shutil.copy2(first, moved / first.name)
        shutil.copy2(second, moved / second.name)
        one = create_master_lineage([first, second])
        two = create_master_lineage([moved / second.name, moved / first.name])
        self.assertEqual(one, two)
        self.assertTrue(one.source_version.startswith("nids-master-v1:"))
        self.assertNotIn(str(self.temp_dir), json.dumps(one.canonical_payload()))

    def test_lineage_changes_when_workbook_content_changes(self) -> None:
        before = create_master_lineage([self.master_path])
        workbook = load_workbook(self.master_path)
        workbook["a-sheet"].append([9, 90, 900, None])
        workbook.save(self.master_path)
        workbook.close()
        after = create_master_lineage([self.master_path])
        self.assertNotEqual(before.source_version, after.source_version)

    def test_integer_code_normalization_contract(self) -> None:
        for value in (" 0010 ", "10.0", 10, Decimal("10.000"), 10.0, np.int64(10)):
            self.assertEqual(normalize_integer_code(value), "10")
        for value in (None, "", True, -1, 1.5, Decimal("1.1"), float("inf"), float(2**53 + 2)):
            self.assertIsNone(normalize_integer_code(value))

    def test_stream_reports_invalid_and_duplicate_keys_with_bounded_diagnostics(self) -> None:
        path = self.temp_dir / "invalid.xlsx"
        rows = [MASTER_HEADERS, [1, 2, 3, None], ["", 2, 3, None]]
        rows.extend([[None, 2, 3, None] for _ in range(25)])
        write_master(path, [("data", rows)])
        stream = stream_master_product_keys([path])
        self.assertEqual(list(stream), [("1", "2", "3")])
        self.assertEqual(stream.report.rows_read, 27)
        self.assertEqual(stream.report.invalid_key_rows, 26)
        self.assertEqual(len(stream.report.invalid_keys.sample), 20)
        self.assertEqual(stream.report.invalid_keys.omitted, 6)

    def test_stream_context_closes_workbook_after_early_exit_and_is_one_pass(self) -> None:
        original = load_workbook(self.master_path, read_only=True, data_only=True)
        proxy = TrackingWorkbook(original)
        with patch.object(lookup_module, "load_workbook", return_value=proxy):
            with stream_master_product_keys([self.master_path]) as stream:
                iterator = iter(stream)
                next(iterator)
            self.assertEqual(proxy.close_calls, 1)
            stream.close()
            with self.assertRaises(RuntimeError):
                iter(stream)

    def test_discovery_and_streaming_error_close_workbook(self) -> None:
        discovery_proxy = TrackingWorkbook(
            load_workbook(self.master_path, read_only=True, data_only=True)
        )
        with patch.object(lookup_module, "load_workbook", return_value=discovery_proxy):
            self.assertEqual(len(discover_master_sheets(self.master_path)), 3)
        self.assertEqual(discovery_proxy.close_calls, 1)

        stream_proxy = TrackingWorkbook(
            load_workbook(self.master_path, read_only=True, data_only=True)
        )
        with (
            patch.object(lookup_module, "load_workbook", return_value=stream_proxy),
            patch.object(lookup_module, "normalize_integer_code", side_effect=RuntimeError("synthetic failure")),
        ):
            with self.assertRaises(lookup_module.MasterProductLookupError):
                list(stream_master_product_keys([self.master_path]))
        self.assertEqual(stream_proxy.close_calls, 1)

    def test_builds_exact_path_schema_manifest_and_deduplicates(self) -> None:
        self.master_path = default_master(self.master_path, duplicate=True)
        result = self.build()
        final = self.final_dir(result.source_hash)
        self.assertEqual(result.status, "written")
        self.assertTrue((final / DATABASE_FILENAME).is_file())
        manifest_bytes = (final / MANIFEST_FILENAME).read_bytes()
        manifest = json.loads(manifest_bytes)
        self.assertEqual(manifest_bytes, lookup_module._canonical_json_bytes(manifest))
        self.assertNotIn(":", final.name)
        self.assertEqual(result.rows_read, 4)
        self.assertEqual(result.unique_key_count, 3)
        self.assertEqual(result.duplicate_key_rows, 1)
        connection = sqlite3.connect(final / DATABASE_FILENAME)
        try:
            schema = connection.execute(
                "SELECT sql FROM sqlite_master WHERE name='product_key'"
            ).fetchone()[0]
            self.assertIn("WITHOUT ROWID", schema.upper())
            self.assertEqual([row[1] for row in connection.execute("PRAGMA table_info(product_key)")], list(lookup_module.PRODUCT_KEY_COLUMNS))
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM product_key").fetchone()[0], 3)
        finally:
            connection.close()

    def test_build_result_and_manifest_keep_bounded_invalid_key_report(self) -> None:
        path = self.temp_dir / "invalid-build.xlsx"
        rows = [MASTER_HEADERS, [1, 2, 3, None]]
        rows.extend([[None, 2, 3, None] for _ in range(25)])
        write_master(path, [("data", rows)])
        result = build_master_product_lookup([path], self.lookup_root)
        self.assertEqual(result.rows_read, 26)
        self.assertEqual(result.valid_key_rows, 1)
        self.assertEqual(result.invalid_key_rows, 25)
        self.assertEqual(len(result.invalid_key_locations), 20)
        self.assertEqual(result.invalid_key_omitted, 5)

    def test_header_only_master_is_not_published(self) -> None:
        path = self.temp_dir / "header-only.xlsx"
        write_master(path, [("data", [["metadata"], MASTER_HEADERS])])
        lineage = create_master_lineage([path])
        with self.assertRaises(EmptyMasterLookupError):
            build_master_product_lookup([path], self.lookup_root)
        self.assertFalse(self.final_dir(lineage.source_hash).exists())
        self.assertEqual(
            list((self.lookup_root / "master_product_lookup").glob(".lookup.tmp-*")),
            [],
        )

    def test_all_invalid_master_rows_are_not_published(self) -> None:
        path = self.temp_dir / "all-invalid.xlsx"
        write_master(
            path,
            [("data", [MASTER_HEADERS, [None, 2, 3], [1, 1.5, 3], [1, 2, None]])],
        )
        lineage = create_master_lineage([path])
        with self.assertRaises(EmptyMasterLookupError):
            build_master_product_lookup([path], self.lookup_root)
        self.assertFalse(self.final_dir(lineage.source_hash).exists())
        staging_root = self.lookup_root / "master_product_lookup"
        self.assertEqual(list(staging_root.glob(".lookup.tmp-*")), [])
        self.assertEqual(list(staging_root.rglob("*.sqlite")), [])

    def test_identical_rerun_is_unchanged_and_does_not_rewrite(self) -> None:
        first = self.build()
        database = self.final_dir(first.source_hash) / DATABASE_FILENAME
        before = database.stat().st_mtime_ns
        second = self.build()
        self.assertEqual(second.status, "unchanged")
        self.assertEqual(database.stat().st_mtime_ns, before)

    def test_manifest_write_failure_cleans_candidate_without_publishing(self) -> None:
        lineage = create_master_lineage([self.master_path])
        with patch.object(Path, "write_bytes", side_effect=OSError("synthetic failure")):
            with self.assertRaises(MasterLookupStorageError):
                self.build()
        self.assertFalse(self.final_dir(lineage.source_hash).exists())
        staging_root = self.lookup_root / "master_product_lookup"
        self.assertEqual(list(staging_root.glob(".lookup.tmp-*")), [])

    def test_staging_directory_creation_oserror_is_typed(self) -> None:
        lineage = create_master_lineage([self.master_path])
        original_mkdir = Path.mkdir

        def fail_staging(path: Path, *args: object, **kwargs: object) -> None:
            if path.name.startswith(".lookup.tmp-"):
                raise OSError("synthetic staging failure")
            original_mkdir(path, *args, **kwargs)

        with patch.object(Path, "mkdir", new=fail_staging):
            with self.assertRaises(MasterLookupStorageError):
                self.build()
        self.assertFalse(self.final_dir(lineage.source_hash).exists())
        self.assertEqual(
            list((self.lookup_root / "master_product_lookup").glob(".lookup.tmp-*")),
            [],
        )

    def test_changed_source_gets_distinct_immutable_path(self) -> None:
        first = self.build()
        changed = self.temp_dir / "changed.xlsx"
        default_master(changed)
        workbook = load_workbook(changed)
        workbook["a-sheet"].append([9, 90, 900, None])
        workbook.save(changed)
        workbook.close()
        second = build_master_product_lookup([changed], self.lookup_root)
        self.assertNotEqual(first.source_hash, second.source_hash)
        self.assertTrue(self.final_dir(first.source_hash).is_dir())
        self.assertTrue(self.final_dir(second.source_hash).is_dir())

    def test_missing_manifest_and_database_corruption_are_rejected(self) -> None:
        result = self.build()
        final = self.final_dir(result.source_hash)
        manifest = final / MANIFEST_FILENAME
        saved = manifest.read_bytes()
        manifest.unlink()
        with self.assertRaises(MasterLookupIntegrityError):
            verify_master_product_lookup(self.lookup_root, result.source_hash)
        manifest.write_bytes(saved)
        with (final / DATABASE_FILENAME).open("ab") as stream:
            stream.write(b"tamper")
        with self.assertRaisesRegex(MasterLookupIntegrityError, "size"):
            verify_master_product_lookup(self.lookup_root, result.source_hash)

    def test_publish_race_identical_is_unchanged_and_different_conflicts(self) -> None:
        first = self.build()
        final = self.final_dir(first.source_hash)
        manifest = json.loads((final / MANIFEST_FILENAME).read_text(encoding="utf-8"))
        candidate = self.temp_dir / "candidate"
        shutil.copytree(final, candidate)
        with patch.object(Path, "replace", side_effect=FileExistsError("race")):
            self.assertFalse(
                lookup_module._publish_candidate_lookup(
                    candidate, final, self.lookup_root, first.source_hash, manifest
                )
            )
        different = dict(manifest)
        different["database_sha256"] = "0" * 64
        with patch.object(Path, "replace", side_effect=FileExistsError("race")):
            with self.assertRaises(MasterLookupConflictError):
                lookup_module._publish_candidate_lookup(
                    candidate, final, self.lookup_root, first.source_hash, different
                )

    def test_verify_returns_source_and_physical_identity(self) -> None:
        result = self.build()
        verified = verify_master_product_lookup(self.lookup_root, result.source_hash)
        self.assertEqual(verified.source_version, result.source_version)
        self.assertEqual(verified.unique_key_count, 3)
        self.assertRegex(verified.database_sha256, r"^[0-9a-f]{64}$")

    def test_join_matches_exact_three_key_preserves_order_and_reports_unmatched(self) -> None:
        result = self.build()
        batch = supply_batch([(3, 30, 300), (99, 30, 300), (1, 10, 100), (2, 20, 200)])
        with open_master_product_lookup(self.lookup_root, result.source_hash) as lookup:
            joined = join_supply_batch_to_master(batch, lookup)
        self.assertEqual(joined.matched_rows.index.tolist(), [0, 2, 3])
        self.assertEqual(joined.report.rows_input, 4)
        self.assertEqual(joined.report.rows_matched, 3)
        self.assertEqual(joined.report.rows_unmatched, 1)
        self.assertLessEqual(joined.report.rows_matched, joined.report.rows_input)
        self.assertEqual(joined.report.match_rate, Decimal("0.75"))
        self.assertEqual(joined.report.unmatched_source_row_ids, (batch.iloc[1]["source_row_id"],))

    def test_same_udi_does_not_match_when_other_two_keys_differ(self) -> None:
        result = self.build()
        batch = supply_batch([(999, 999, 100)])
        with open_master_product_lookup(self.lookup_root, result.source_hash) as lookup:
            joined = lookup.join_supply_batch(batch)
        self.assertTrue(joined.matched_rows.empty)
        self.assertEqual(joined.report.rows_unmatched, 1)

    def test_duplicate_supply_rows_are_preserved_and_database_is_read_only(self) -> None:
        result = self.build()
        batch = supply_batch([(1, 10, 100), (1, 10, 100)])
        with open_master_product_lookup(self.lookup_root, result.source_hash) as lookup:
            joined = lookup.join_supply_batch(batch)
            with self.assertRaises(sqlite3.OperationalError):
                lookup._connection.execute(
                    "INSERT INTO product_key VALUES ('9', '9', '9')"
                )
        self.assertEqual(joined.matched_rows.index.tolist(), [0, 1])
        self.assertEqual(joined.report.rows_matched, 2)

    def test_repeated_joins_leave_no_transaction_or_temp_table(self) -> None:
        result = self.build()
        batch = supply_batch([(1, 10, 100), (99, 99, 99)])
        with open_master_product_lookup(self.lookup_root, result.source_hash) as lookup:
            for _ in range(3):
                joined = lookup.join_supply_batch(batch)
                self.assertEqual(joined.report.rows_matched, 1)
                self.assertFalse(lookup._connection.in_transaction)
                temp_tables = lookup._connection.execute(
                    "SELECT name FROM sqlite_temp_master "
                    "WHERE type='table' AND name LIKE 'batch_product_key_%'"
                ).fetchall()
                self.assertEqual(temp_tables, [])

    def test_join_sql_error_rolls_back_and_removes_temp_table(self) -> None:
        result = self.build()
        batch = supply_batch([(1, 10, 100)])
        with open_master_product_lookup(self.lookup_root, result.source_hash) as lookup:
            with patch.object(
                lookup_module,
                "_select_matching_batch_positions",
                side_effect=sqlite3.OperationalError("synthetic join failure"),
            ):
                with self.assertRaises(SupplyBatchJoinError):
                    lookup.join_supply_batch(batch)
            self.assertFalse(lookup._connection.in_transaction)
            temp_tables = lookup._connection.execute(
                "SELECT name FROM sqlite_temp_master "
                "WHERE type='table' AND name LIKE 'batch_product_key_%'"
            ).fetchall()
            self.assertEqual(temp_tables, [])
            self.assertEqual(lookup.join_supply_batch(batch).report.rows_matched, 1)

    def test_join_rejects_non_exact_schema_multiple_versions_and_incomplete_keys(self) -> None:
        result = self.build()
        with open_master_product_lookup(self.lookup_root, result.source_hash) as lookup:
            with self.assertRaises(SupplyBatchJoinError):
                lookup.join_supply_batch(supply_batch([(1, 10, 100)]).drop(columns=["udi"]))
            mixed = supply_batch([(1, 10, 100), (2, 20, 200)])
            mixed.loc[1, "source_version"] = "other"
            with self.assertRaisesRegex(SupplyBatchJoinError, "one source_version"):
                lookup.join_supply_batch(mixed)
            incomplete = supply_batch([(1, None, 100)])
            with self.assertRaisesRegex(SupplyBatchJoinError, "incomplete"):
                lookup.join_supply_batch(incomplete)

    def test_unmatched_diagnostics_are_bounded_to_twenty(self) -> None:
        result = self.build()
        batch = supply_batch([(1000 + number, 2000 + number, 3000 + number) for number in range(25)])
        with open_master_product_lookup(self.lookup_root, result.source_hash) as lookup:
            report = lookup.join_supply_batch(batch).report
        self.assertEqual(report.rows_unmatched, 25)
        self.assertEqual(len(report.unmatched_source_row_ids), 20)
        self.assertEqual(report.unmatched_omitted, 5)

    def test_empty_batch_is_rejected_without_inventing_a_supply_version(self) -> None:
        result = self.build()
        batch = supply_batch([])
        with open_master_product_lookup(self.lookup_root, result.source_hash) as lookup:
            with self.assertRaisesRegex(SupplyBatchJoinError, "one source_version"):
                lookup.join_supply_batch(batch)

    def test_close_releases_database_for_windows_rename_and_is_idempotent(self) -> None:
        result = self.build()
        final = self.final_dir(result.source_hash)
        database = final / DATABASE_FILENAME
        lookup = open_master_product_lookup(self.lookup_root, result.source_hash)
        lookup.close()
        lookup.close()
        moved = final / "renamed.sqlite"
        database.replace(moved)
        moved.replace(database)

    def test_build_does_not_store_non_key_master_values(self) -> None:
        result = self.build()
        database = self.final_dir(result.source_hash) / DATABASE_FILENAME
        connection = sqlite3.connect(database)
        try:
            tables = connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
            self.assertEqual(tables, [("product_key",)])
            self.assertNotIn("not stored", database.read_bytes().decode("latin1"))
        finally:
            connection.close()


if __name__ == "__main__":
    unittest.main()
