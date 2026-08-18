from __future__ import annotations

from decimal import Decimal
from pathlib import Path
import shutil
import unittest
from unittest.mock import patch
from uuid import uuid4

import pandas as pd
from pandas.testing import assert_frame_equal
from openpyxl import Workbook

from data_pipeline.contracts import SOURCE_REQUIRED_COLUMNS, normalize_source_rows
from data_pipeline.ingest import nids_supply_excel as adapter
from data_pipeline.ingest.nids_supply_excel import (
    ClosedSupplyMonth,
    DataSheetDiscoveryError,
    DataSheetSchemaError,
    NidsSupplyExcelError,
    RejectedSupplyMonth,
    SOURCE_BATCH_COLUMNS,
    SupplyWorkbookNameError,
    create_source_lineage,
    declared_month_from_logical_names,
    discover_supply_sheets,
    group_closed_supply_months,
    parse_supply_workbook_date_range,
    stream_nids_supply_excel,
)


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
TEMP_PARENT = REPOSITORY_ROOT.parent

HEADERS = [
    "공급일자",
    "공급한자 업체일련번호",
    "공급받은자 업체일련번호",
    "요양기관기호(의료기관)",
    "의료기기품목일련번호",
    "모델일련번호",
    "UDI-DI 일련번호",
    "품목군",
    "품목명",
    "공급구분",
    "공급금액",
    "공급수량",
    "포장내 총 수량",
    "낱개총수량",
    "UDI-DI",
    "업종",
    "공급받은자업종",
    "공급한자의 소재지 시도코드",
    "공급받은자의 소재지 시도코드",
    "거래처 코드",
    "공급내역기준연월",
    "공급내역작업일련번호",
    "공급내역일련번호",
    "공급내역보고자료복합Key",
    "사용하지 않는 열",
]


def row(**overrides: object) -> list[object]:
    values: dict[str, object] = {
        "공급일자": "20260115",
        "공급한자 업체일련번호": " 10 ",
        "공급받은자 업체일련번호": "20",
        "요양기관기호(의료기관)": None,
        "의료기기품목일련번호": "001",
        "모델일련번호": 2,
        "UDI-DI 일련번호": "3.0",
        "품목군": "SYNTHETIC-GROUP",
        "품목명": "SYNTHETIC-ITEM",
        "공급구분": "출고",
        "공급금액": "1234567890.123456",
        "공급수량": "2",
        "포장내 총 수량": "5",
        "낱개총수량": "10",
        "UDI-DI": "000012345",
        "업종": "SYNTHETIC-SUPPLIER",
        "공급받은자업종": "SYNTHETIC-RECEIVER",
        "공급한자의 소재지 시도코드": "01",
        "공급받은자의 소재지 시도코드": "02",
        "거래처 코드": "100",
        "공급내역기준연월": "202601",
        "공급내역작업일련번호": "200",
        "공급내역일련번호": "300",
        "공급내역보고자료복합Key": "comparison-only",
        "사용하지 않는 열": "not-emitted",
    }
    values.update(overrides)
    return [values.get(header) for header in HEADERS]


def row_for_headers(headers: list[str], **overrides: object) -> list[object]:
    values = dict(zip(HEADERS, row(**overrides)))
    return [values.get(header) for header in headers]


def write_workbook(
    path: Path,
    *,
    sheets: list[tuple[str, list[list[object]]]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets:
        sheet = workbook.create_sheet(name)
        for values in rows:
            sheet.append(values)
    workbook.save(path)
    workbook.close()


def data_rows(path: Path, *, batch_size: int = 100) -> tuple[pd.DataFrame, object]:
    stream = stream_nids_supply_excel([path], batch_size=batch_size)
    batches = list(stream)
    frame = (
        pd.concat(batches, ignore_index=True)
        if batches
        else pd.DataFrame(columns=SOURCE_BATCH_COLUMNS)
    )
    return frame, stream.report


class TrackingWorkbook:
    def __init__(self, workbook: object) -> None:
        self.workbook = workbook
        self.close_calls = 0

    @property
    def sheetnames(self) -> list[str]:
        return self.workbook.sheetnames

    def __getitem__(self, name: str) -> object:
        return self.workbook[name]

    def close(self) -> None:
        self.close_calls += 1
        self.workbook.close()


class NidsSupplyExcelAdapterTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TEMP_PARENT / f".nids-ingest-{uuid4().hex[:10]}"
        self.temp_dir.mkdir(parents=True)

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def workbook(self, name: str = "synthetic.xlsx") -> Path:
        path = self.temp_dir / name
        write_workbook(
            path,
            sheets=[
                ("arbitrary metadata", [["synthetic overview"], ["not data"]]),
                ("random data title", [["intro"], HEADERS, row()]),
            ],
        )
        return path

    def test_discovers_arbitrary_metadata_and_data_names_with_shifted_header(self) -> None:
        discovered = discover_supply_sheets(self.workbook(), header_scan_limit=5)

        self.assertEqual(len(discovered), 1)
        self.assertEqual(discovered[0].name, "random data title")
        self.assertEqual(discovered[0].header_row, 2)

    def test_discovery_order_is_deterministic_when_sheet_order_changes(self) -> None:
        first = self.temp_dir / "first.xlsx"
        second = self.temp_dir / "second.xlsx"
        data_a = ("z-data", [HEADERS, row(**{"공급내역일련번호": "301"})])
        data_b = ("a-data", [HEADERS, row(**{"공급내역일련번호": "302"})])
        write_workbook(first, sheets=[data_a, data_b])
        write_workbook(second, sheets=[data_b, data_a])

        self.assertEqual(
            tuple(sheet.name for sheet in discover_supply_sheets(first)),
            ("a-data", "z-data"),
        )
        self.assertEqual(
            tuple(sheet.name for sheet in discover_supply_sheets(second)),
            ("a-data", "z-data"),
        )

    def test_reads_all_discovered_data_sheets(self) -> None:
        path = self.temp_dir / "multi.xlsx"
        write_workbook(
            path,
            sheets=[
                ("one", [HEADERS, row(**{"공급내역일련번호": "301"})]),
                ("metadata", [["nothing"]]),
                ("two", [HEADERS, row(**{"공급내역일련번호": "302"})]),
            ],
        )

        actual, report = data_rows(path)

        self.assertEqual(len(actual), 2)
        self.assertEqual(report.rows_read, 2)
        self.assertEqual(report.rows_emitted, 2)

    def test_no_data_sheet_and_ambiguous_header_are_explicit_errors(self) -> None:
        no_data = self.temp_dir / "no-data.xlsx"
        write_workbook(no_data, sheets=[("metadata", [["nothing"]])])
        with self.assertRaises(DataSheetDiscoveryError):
            discover_supply_sheets(no_data)

        ambiguous = self.temp_dir / "ambiguous.xlsx"
        write_workbook(ambiguous, sheets=[("data", [HEADERS, HEADERS, row()])])
        with self.assertRaisesRegex(DataSheetDiscoveryError, "Ambiguous"):
            discover_supply_sheets(ambiguous)

    def test_duplicate_headers_are_rejected(self) -> None:
        path = self.temp_dir / "duplicate.xlsx"
        duplicate_headers = HEADERS + ["공급일자"]
        write_workbook(path, sheets=[("data", [duplicate_headers])])
        with self.assertRaisesRegex(DataSheetDiscoveryError, "Duplicate"):
            discover_supply_sheets(path)

    def test_missing_required_mapped_fields_fail_before_row_mapping(self) -> None:
        for missing_header, logical_field in (
            ("모델일련번호", "model_serial"),
            ("UDI-DI 일련번호", "udi_serial"),
        ):
            with self.subTest(logical_field=logical_field):
                headers = [header for header in HEADERS if header != missing_header]
                path = self.temp_dir / f"missing-{logical_field}.xlsx"
                write_workbook(
                    path,
                    sheets=[("synthetic-data", [headers, row_for_headers(headers)])],
                )
                stream = stream_nids_supply_excel([path])
                with patch.object(adapter, "_map_row") as mapped:
                    with self.assertRaisesRegex(
                        DataSheetSchemaError, logical_field
                    ) as raised:
                        list(stream)
                mapped.assert_not_called()
                message = str(raised.exception)
                self.assertIn(path.name, message)
                self.assertIn("synthetic-data", message)
                self.assertNotIn("SYNTHETIC-ITEM", message)

    def test_both_receiver_columns_missing_fail_before_row_mapping(self) -> None:
        headers = [
            header
            for header in HEADERS
            if header
            not in {"공급받은자 업체일련번호", "요양기관기호(의료기관)"}
        ]
        path = self.temp_dir / "missing-receiver-columns.xlsx"
        write_workbook(
            path,
            sheets=[("data", [headers, row_for_headers(headers)])],
        )
        stream = stream_nids_supply_excel([path])

        with patch.object(adapter, "_map_row") as mapped:
            with self.assertRaisesRegex(
                DataSheetSchemaError, r"dst_company_id\|hospital_id"
            ):
                list(stream)
        mapped.assert_not_called()

    def test_hospital_column_alone_satisfies_receiver_structure(self) -> None:
        headers = [
            header for header in HEADERS if header != "공급받은자 업체일련번호"
        ]
        path = self.temp_dir / "hospital-only.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [
                        headers,
                        row_for_headers(
                            headers,
                            **{"요양기관기호(의료기관)": "001234"},
                        ),
                    ],
                )
            ],
        )

        actual, report = data_rows(path)

        self.assertEqual(len(actual), 1)
        self.assertEqual(actual.loc[0, "dst_company_id"], "hosp:001234")
        self.assertIn(
            "공급받은자 업체일련번호",
            report.sheet_profiles[0].missing_columns,
        )

    def test_optional_columns_missing_are_profiled_but_streaming_continues(self) -> None:
        optional_headers = {
            "공급금액",
            "포장내 총 수량",
            "낱개총수량",
            "품목군",
            "품목명",
            "UDI-DI",
            "업종",
            "공급받은자업종",
            "공급한자의 소재지 시도코드",
            "공급받은자의 소재지 시도코드",
            "공급내역보고자료복합Key",
        }
        headers = [header for header in HEADERS if header not in optional_headers]
        path = self.temp_dir / "optional-missing.xlsx"
        write_workbook(
            path,
            sheets=[("data", [headers, row_for_headers(headers)])],
        )

        actual, report = data_rows(path)

        self.assertEqual(len(actual), 1)
        missing = set(report.sheet_profiles[0].missing_columns)
        self.assertTrue(optional_headers.issubset(missing))
        self.assertTrue(pd.isna(actual.loc[0, "amount_clean"]))
        self.assertTrue(pd.isna(actual.loc[0, "piece_qty"]))

    def test_batch_size_is_a_hard_upper_bound(self) -> None:
        path = self.temp_dir / "batches.xlsx"
        rows = [HEADERS] + [
            row(**{"공급내역일련번호": str(300 + index)}) for index in range(7)
        ]
        write_workbook(path, sheets=[("data", rows)])
        stream = stream_nids_supply_excel([path], batch_size=3)

        batches = list(stream)

        self.assertEqual([len(batch) for batch in batches], [3, 3, 1])
        self.assertTrue(all(len(batch) <= 3 for batch in batches))

    def test_source_version_ignores_input_order_and_absolute_location(self) -> None:
        left = self.temp_dir / "left"
        right = self.temp_dir / "right"
        left.mkdir()
        right.mkdir()
        a = left / "a.xlsx"
        b = left / "b.xlsx"
        write_workbook(a, sheets=[("data", [HEADERS, row()])])
        write_workbook(b, sheets=[("data", [HEADERS, row()])])
        shutil.copy2(a, right / a.name)
        shutil.copy2(b, right / b.name)

        first = create_source_lineage([a, b])
        second = create_source_lineage([right / b.name, right / a.name])

        self.assertEqual(first.source_version, second.source_version)
        self.assertNotIn(str(left), str(first.canonical_payload()))

    def test_source_version_changes_when_file_content_changes(self) -> None:
        path = self.workbook()
        before = create_source_lineage([path]).source_version
        write_workbook(
            path,
            sheets=[("data", [HEADERS, row(**{"공급금액": "999"})])],
        )
        after = create_source_lineage([path]).source_version
        self.assertNotEqual(before, after)

    def test_source_row_id_is_deterministic_and_uses_normalized_components(self) -> None:
        first = self.workbook("first-row.xlsx")
        second = self.temp_dir / "second-row.xlsx"
        write_workbook(
            second,
            sheets=[
                (
                    "data",
                    [
                        HEADERS,
                        row(
                            **{
                                "거래처 코드": 100.0,
                                "공급내역기준연월": 202601,
                                "공급내역작업일련번호": Decimal("200"),
                                "공급내역일련번호": "0300.0",
                            }
                        ),
                    ],
                )
            ],
        )

        first_rows, _ = data_rows(first)
        second_rows, _ = data_rows(second)
        self.assertEqual(first_rows.loc[0, "source_row_id"], second_rows.loc[0, "source_row_id"])
        self.assertEqual(
            first_rows.loc[0, "source_row_id"],
            "nids-row-v1:e944f43a81c2c89d0a070a399eab062be34514072a894c21a1553d7abe0ce43b",
        )

    def test_incomplete_source_identity_is_blocked_with_bounded_report(self) -> None:
        path = self.temp_dir / "blocked.xlsx"
        rows = [HEADERS] + [
            row(
                **{
                    "공급내역일련번호": None,
                    "공급내역작업일련번호": str(index),
                }
            )
            for index in range(25)
        ]
        write_workbook(path, sheets=[("data", rows)])

        actual, report = data_rows(path)

        self.assertTrue(actual.empty)
        issue = report.source_identity_incomplete
        self.assertEqual(issue.total, 25)
        self.assertEqual(issue.status, "blocked:deduplication_unverified")
        self.assertEqual(len(issue.sample), 20)
        self.assertEqual(issue.omitted, 5)
        self.assertTrue(all("row=" in sample for sample in issue.sample))
        self.assertEqual(report.rows_read, 25)
        self.assertEqual(report.rows_emitted, 0)
        self.assertEqual(report.rows_rejected, 25)
        self.assertEqual(
            report.rejected_by_reason,
            {"source_identity_incomplete": 25},
        )
        self.assertTrue(report.accounting_is_complete)

    def test_exclusive_rejection_accounting_across_multiple_sheets(self) -> None:
        path = self.temp_dir / "accounting.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "alpha",
                    [
                        HEADERS,
                        row(),
                        row(
                            **{
                                "공급내역일련번호": None,
                                "공급한자 업체일련번호": None,
                                "모델일련번호": None,
                                "공급일자": "invalid",
                            }
                        ),
                        row(
                            **{
                                "공급내역일련번호": "301",
                                "공급한자 업체일련번호": None,
                            }
                        ),
                    ],
                ),
                (
                    "beta",
                    [
                        HEADERS,
                        row(
                            **{
                                "공급내역일련번호": "302",
                                "모델일련번호": None,
                            }
                        ),
                        row(
                            **{
                                "공급내역일련번호": "303",
                                "공급일자": "invalid",
                            }
                        ),
                        row(**{"공급내역일련번호": "304"}),
                    ],
                ),
            ],
        )

        actual, report = data_rows(path, batch_size=2)

        self.assertEqual(len(actual), 2)
        self.assertEqual(report.rows_read, 6)
        self.assertEqual(report.rows_emitted, 2)
        self.assertEqual(report.rows_rejected, 4)
        self.assertEqual(
            report.rejected_by_reason,
            {
                "source_identity_incomplete": 1,
                "party_identity_incomplete": 1,
                "product_key_incomplete": 1,
                "date_invalid": 1,
            },
        )
        self.assertEqual(
            report.rows_read,
            report.rows_emitted + report.rows_rejected,
        )
        self.assertEqual(
            report.rows_rejected,
            sum(report.rejected_by_reason.values()),
        )
        self.assertTrue(report.accounting_is_complete)
        report.validate_accounting()

    def test_rejection_report_memory_is_bounded_by_reason_and_sample_limit(self) -> None:
        path = self.temp_dir / "bounded-rejections.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [HEADERS]
                    + [
                        row(
                            **{
                                "공급내역작업일련번호": str(index),
                                "공급내역일련번호": None,
                            }
                        )
                        for index in range(1_000)
                    ],
                )
            ],
        )

        actual, report = data_rows(path)

        self.assertTrue(actual.empty)
        self.assertEqual(report.rows_rejected, 1_000)
        self.assertEqual(len(report.rejected_by_reason), 1)
        self.assertEqual(len(report.source_identity_incomplete.sample), 20)
        self.assertEqual(report.source_identity_incomplete.omitted, 980)
        self.assertTrue(report.accounting_is_complete)

    def test_reported_composite_key_is_quality_only_and_never_an_id_fallback(self) -> None:
        path = self.temp_dir / "composite.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [
                        HEADERS,
                        row(
                            **{
                                "공급내역일련번호": None,
                                "공급내역보고자료복합Key": "present-but-not-a-fallback",
                            }
                        ),
                        row(
                            **{
                                "공급내역일련번호": "301",
                                "공급내역보고자료복합Key": None,
                            }
                        ),
                    ],
                )
            ],
        )

        actual, report = data_rows(path)

        self.assertEqual(len(actual), 1)
        self.assertEqual(report.source_identity_incomplete.total, 1)
        self.assertEqual(report.reported_composite_key_available, 1)
        self.assertEqual(report.reported_composite_key_missing.total, 1)
        self.assertEqual(
            report.reported_composite_key_present_identity_incomplete.total,
            1,
        )

    def test_invalid_date_and_negative_values_are_not_normalized_as_valid(self) -> None:
        path = self.temp_dir / "invalid-values.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [
                        HEADERS,
                        row(**{"공급일자": "not-a-date"}),
                        row(
                            **{
                                "공급내역일련번호": "301",
                                "공급금액": "-1",
                                "공급수량": "-2",
                                "낱개총수량": "-10",
                            }
                        ),
                    ],
                )
            ],
        )

        actual, report = data_rows(path)

        self.assertEqual(len(actual), 1)
        self.assertEqual(report.date_conversion_failed.total, 1)
        self.assertEqual(report.amount_conversion_failed.total, 1)
        self.assertEqual(report.quantity_conversion_failed.total, 2)
        self.assertTrue(pd.isna(actual.loc[0, "amount_clean"]))
        self.assertTrue(pd.isna(actual.loc[0, "raw_supply_qty"]))
        self.assertTrue(pd.isna(actual.loc[0, "piece_qty"]))

    def test_party_fallback_and_missing_party_rules(self) -> None:
        path = self.temp_dir / "parties.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [
                        HEADERS,
                        row(
                            **{
                                "공급받은자 업체일련번호": None,
                                "요양기관기호(의료기관)": "001234",
                            }
                        ),
                        row(
                            **{
                                "공급내역일련번호": "301",
                                "공급받은자 업체일련번호": None,
                                "요양기관기호(의료기관)": None,
                            }
                        ),
                    ],
                )
            ],
        )

        actual, report = data_rows(path)

        self.assertEqual(len(actual), 1)
        self.assertEqual(actual.loc[0, "dst_company_id"], "hosp:001234")
        self.assertIn("receiver_hospital_fallback", actual.loc[0, "row_quality_flags"])
        self.assertEqual(report.party_identity_incomplete.total, 1)

    def test_incomplete_three_key_is_blocked(self) -> None:
        path = self.temp_dir / "product-key.xlsx"
        write_workbook(
            path,
            sheets=[("data", [HEADERS, row(**{"모델일련번호": None})])],
        )

        actual, report = data_rows(path)

        self.assertTrue(actual.empty)
        self.assertEqual(report.product_key_incomplete.total, 1)

    def test_decimal_precision_is_preserved_without_float_conversion(self) -> None:
        actual, _ = data_rows(self.workbook())

        self.assertIsInstance(actual.loc[0, "amount_clean"], Decimal)
        self.assertEqual(actual.loc[0, "amount_clean"], Decimal("1234567890.123456"))
        self.assertEqual(actual.loc[0, "raw_supply_qty"], Decimal("2"))

    def test_transaction_types_are_mapped_without_unknown_supply_fallback(self) -> None:
        path = self.temp_dir / "types.xlsx"
        types = ["출고", "반품", "회수", "폐기", "임대", "기타"]
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [HEADERS]
                    + [
                        row(
                            **{
                                "공급구분": value,
                                "공급내역일련번호": str(300 + index),
                            }
                        )
                        for index, value in enumerate(types)
                    ],
                )
            ],
        )

        actual, report = data_rows(path)

        self.assertEqual(
            list(actual["transaction_type"]),
            ["SUPPLY", "RETURN", "RECALL", "DISCARD", "LEASE", "기타"],
        )
        self.assertEqual(report.transaction_type_unknown.total, 1)
        self.assertNotEqual(actual.iloc[-1]["transaction_type"], "SUPPLY")

    def test_piece_quantity_valid_mismatch_and_unverified(self) -> None:
        path = self.temp_dir / "pieces.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [
                        HEADERS,
                        row(),
                        row(
                            **{
                                "공급내역일련번호": "301",
                                "낱개총수량": "11",
                            }
                        ),
                        row(
                            **{
                                "공급내역일련번호": "302",
                                "포장내 총 수량": None,
                                "낱개총수량": "10",
                            }
                        ),
                    ],
                )
            ],
        )

        actual, report = data_rows(path)

        self.assertEqual(actual.loc[0, "piece_qty"], Decimal("10"))
        self.assertTrue(pd.isna(actual.loc[1, "piece_qty"]))
        self.assertTrue(pd.isna(actual.loc[2, "piece_qty"]))
        self.assertEqual(report.piece_quantity_mismatch.total, 1)
        self.assertEqual(report.piece_quantity_unverified.total, 1)

    def test_extreme_amounts_are_preserved_and_profiled(self) -> None:
        path = self.temp_dir / "amounts.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [
                        HEADERS,
                        row(**{"공급금액": "50000001"}),
                        row(
                            **{
                                "공급내역일련번호": "301",
                                "공급금액": "1000000000001",
                            }
                        ),
                    ],
                )
            ],
        )

        actual, report = data_rows(path)

        self.assertEqual(actual.loc[1, "amount_clean"], Decimal("1000000000001"))
        self.assertEqual(report.high_value_review.total, 2)
        self.assertEqual(report.barcode_entry_error_suspected.total, 1)
        self.assertEqual(report.high_value_max, Decimal("1000000000001"))

    def test_context_exit_after_first_batch_closes_active_workbook(self) -> None:
        path = self.temp_dir / "early-exit.xlsx"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [
                        HEADERS,
                        row(),
                        row(**{"공급내역일련번호": "301"}),
                    ],
                )
            ],
        )
        stream = stream_nids_supply_excel([path], batch_size=1)
        proxy = TrackingWorkbook(
            adapter.load_workbook(path, read_only=True, data_only=True)
        )

        with patch.object(adapter, "load_workbook", return_value=proxy):
            with stream:
                for batch in stream:
                    self.assertEqual(len(batch), 1)
                    break

        self.assertEqual(proxy.close_calls, 1)
        with self.assertRaises(RuntimeError):
            iter(stream)

    def test_streaming_exception_closes_workbook(self) -> None:
        path = self.workbook("stream-error.xlsx")
        stream = stream_nids_supply_excel([path])
        proxy = TrackingWorkbook(
            adapter.load_workbook(path, read_only=True, data_only=True)
        )

        with (
            patch.object(adapter, "load_workbook", return_value=proxy),
            patch.object(adapter, "_map_row", side_effect=RuntimeError("synthetic")),
        ):
            with self.assertRaises(NidsSupplyExcelError):
                list(stream)

        self.assertEqual(proxy.close_calls, 1)

    def test_full_stream_closes_workbook_and_remains_one_pass(self) -> None:
        path = self.workbook("full-close.xlsx")
        stream = stream_nids_supply_excel([path])
        proxy = TrackingWorkbook(
            adapter.load_workbook(path, read_only=True, data_only=True)
        )

        with patch.object(adapter, "load_workbook", return_value=proxy):
            batches = list(stream)

        self.assertEqual(sum(len(batch) for batch in batches), 1)
        self.assertEqual(proxy.close_calls, 1)
        with self.assertRaises(RuntimeError):
            list(stream)

    def test_close_is_idempotent_while_generator_is_active(self) -> None:
        path = self.workbook("double-close.xlsx")
        stream = stream_nids_supply_excel([path], batch_size=1)
        proxy = TrackingWorkbook(
            adapter.load_workbook(path, read_only=True, data_only=True)
        )

        with patch.object(adapter, "load_workbook", return_value=proxy):
            iterator = iter(stream)
            self.assertEqual(len(next(iterator)), 1)
            stream.close()
            stream.close()

        self.assertEqual(proxy.close_calls, 1)
        with self.assertRaises(RuntimeError):
            iter(stream)

    def test_workbook_is_closed_when_discovery_fails(self) -> None:
        path = self.temp_dir / "close.xlsx"
        write_workbook(path, sheets=[("metadata", [["nothing"]])])
        real = adapter.load_workbook(path, read_only=True, data_only=True)

        proxy = TrackingWorkbook(real)
        with patch.object(adapter, "load_workbook", return_value=proxy) as mocked_load:
            with self.assertRaises(DataSheetDiscoveryError):
                discover_supply_sheets(path)
        mocked_load.assert_called_once_with(path, read_only=True, data_only=True)
        self.assertEqual(proxy.close_calls, 1)

    def test_company_display_names_are_not_identifiers_or_batch_columns(self) -> None:
        headers = ["공급자", "공급받은자", *HEADERS]
        path = self.temp_dir / "display-names.xlsx"
        first = dict(zip(HEADERS, row(**{"공급내역일련번호": "301", "공급한자 업체일련번호": "10", "공급받은자 업체일련번호": "20"})))
        first["공급자"] = "알파의료"
        first["공급받은자"] = "베타병원"
        second = dict(zip(HEADERS, row(**{"공급내역일련번호": "302", "공급한자 업체일련번호": "10", "공급받은자 업체일련번호": "20"})))
        second["공급자"] = "알파의료"
        second["공급받은자"] = "다른상호"
        fourth = dict(zip(HEADERS, row(**{"공급내역일련번호": "304", "공급한자 업체일련번호": "10", "공급받은자 업체일련번호": "20"})))
        fourth["공급자"] = "알파의료"
        fourth["공급받은자"] = "베타병원"
        third = dict(zip(HEADERS, row(**{"공급내역일련번호": "303", "공급한자 업체일련번호": None, "공급받은자 업체일련번호": "20"})))
        third["공급자"] = "이름만있는행"
        third["공급받은자"] = "베타병원"
        write_workbook(
            path,
            sheets=[
                (
                    "data",
                    [
                        headers,
                        [first.get(header) for header in headers],
                        [second.get(header) for header in headers],
                        [fourth.get(header) for header in headers],
                        [third.get(header) for header in headers],
                    ],
                )
            ],
        )
        with stream_nids_supply_excel([path]) as stream:
            frame = pd.concat(list(stream), ignore_index=True)
            names = {row["entity_id"]: row for row in stream.display_name_rows()}
            report = stream.report
        self.assertEqual(set(frame["src_company_id"]), {"co:10"})
        self.assertEqual(set(frame["dst_company_id"]), {"co:20"})
        self.assertNotIn("display_name", frame.columns)
        self.assertNotIn("supplier_display_name", frame.columns)
        self.assertNotIn("알파의료", "".join(frame.astype(str).to_numpy().ravel()))
        self.assertNotIn("공급자", report.sheet_profiles[0].extra_columns)
        self.assertEqual(names["co:10"]["display_name"], "알파의료")
        self.assertEqual(names["co:20"]["display_name"], "베타병원")
        self.assertTrue(names["co:20"]["name_conflict"])
        self.assertFalse(names["co:10"]["name_conflict"])
        self.assertEqual(report.rejected_by_reason["party_identity_incomplete"], 1)

    def test_output_columns_match_pr01_source_contract_exactly(self) -> None:
        actual, report = data_rows(self.workbook())

        self.assertEqual(tuple(actual.columns), (*SOURCE_REQUIRED_COLUMNS, "row_quality_flags"))
        self.assertEqual(tuple(actual.columns), SOURCE_BATCH_COLUMNS)
        self.assertEqual(report.sheet_profiles[0].missing_columns, ())
        self.assertIn("사용하지 않는 열", report.sheet_profiles[0].extra_columns)
        self.assertNotIn("공급내역보고자료복합Key", actual.columns)
        normalized = normalize_source_rows(actual)
        self.assertEqual(len(normalized), 1)

    def test_rereading_same_source_produces_identical_rows_and_ids(self) -> None:
        path = self.workbook()

        first, _ = data_rows(path)
        second, _ = data_rows(path)

        assert_frame_equal(first, second)
        self.assertEqual(first.loc[0, "source_row_id"], second.loc[0, "source_row_id"])
        self.assertEqual(first.loc[0, "source_version"], second.loc[0, "source_version"])

    def test_float_decimal_is_not_silently_accepted(self) -> None:
        path = self.temp_dir / "float.xlsx"
        write_workbook(
            path,
            sheets=[("data", [HEADERS, row(**{"공급금액": 1.25})])],
        )

        actual, report = data_rows(path)

        self.assertTrue(pd.isna(actual.loc[0, "amount_clean"]))
        self.assertEqual(report.amount_conversion_failed.total, 1)
        self.assertIn("amount_invalid", actual.loc[0, "row_quality_flags"])


class SupplyWorkbookDateRangeTests(unittest.TestCase):
    def test_parses_real_filename_without_spaces(self) -> None:
        parsed = parse_supply_workbook_date_range("공급내역보고자료(20260101~20260110).xlsx")
        self.assertEqual(parsed.month, "202601")
        self.assertEqual(parsed.start.isoformat(), "2026-01-01")
        self.assertEqual(parsed.end.isoformat(), "2026-01-10")

    def test_february_end_uses_datetime_leap_day(self) -> None:
        parsed = parse_supply_workbook_date_range("공급내역보고자료(20240221~20240229).xlsx")
        self.assertEqual(parsed.month, "202402")
        self.assertEqual(parsed.end.isoformat(), "2024-02-29")

    def test_rejects_non_leap_february_29(self) -> None:
        with self.assertRaises(SupplyWorkbookNameError):
            parse_supply_workbook_date_range("공급내역보고자료(20250221~20250229).xlsx")

    def test_rejects_range_that_crosses_months(self) -> None:
        with self.assertRaises(SupplyWorkbookNameError):
            parse_supply_workbook_date_range("공급내역보고자료(20260121~20260203).xlsx")

    def test_group_closes_exactly_three_files_and_rejects_other_counts(self) -> None:
        root = TEMP_PARENT / f".dekade-group-{uuid4().hex[:8]}"
        root.mkdir()
        try:
            january = [
                root / "공급내역보고자료(20260101~20260110).xlsx",
                root / "공급내역보고자료(20260111~20260120).xlsx",
                root / "공급내역보고자료(20260121~20260131).xlsx",
            ]
            february = [
                root / "공급내역보고자료(20260201~20260210).xlsx",
                root / "공급내역보고자료(20260211~20260220).xlsx",
            ]
            for path in (*january, *february):
                path.write_bytes(b"x")
            grouped = group_closed_supply_months((*january, *february))
            self.assertEqual(len(grouped.closed), 1)
            self.assertEqual(grouped.closed[0].month, "202601")
            self.assertEqual(grouped.closed[0].paths, tuple(sorted(january, key=lambda item: item.name)))
            self.assertEqual(grouped.rejected, (
                RejectedSupplyMonth(
                    "202602", 2,
                    (
                        "공급내역보고자료(20260201~20260210).xlsx",
                        "공급내역보고자료(20260211~20260220).xlsx",
                    ),
                    "expected_exactly_three_files",
                ),
            ))
            self.assertEqual(
                declared_month_from_logical_names([path.name for path in january]),
                "202601",
            )
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def test_four_files_in_one_month_are_rejected(self) -> None:
        root = TEMP_PARENT / f".dekade-four-{uuid4().hex[:8]}"
        root.mkdir()
        try:
            paths = [
                root / "공급내역보고자료(20260401~20260410).xlsx",
                root / "공급내역보고자료(20260411~20260420).xlsx",
                root / "공급내역보고자료(20260421~20260430).xlsx",
                root / "공급내역보고자료(20260421~20260422).xlsx",
            ]
            for path in paths:
                path.write_bytes(b"x")
            grouped = group_closed_supply_months(paths)
            self.assertEqual(grouped.closed, ())
            self.assertEqual(grouped.rejected[0].path_count, 4)
        finally:
            shutil.rmtree(root, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
