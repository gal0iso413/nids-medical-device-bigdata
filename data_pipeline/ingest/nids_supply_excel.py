"""Bounded, content-discovered streaming adapter for NIDS supply workbooks.

This adapter stops at normalized PR-01 source rows and source lineage. It does
not join the master, aggregate monthly facts, or publish Parquet partitions.
"""

from __future__ import annotations

from calendar import monthrange
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from itertools import islice
import json
from pathlib import Path
import re
from typing import Any, Final, Iterable, Iterator, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from data_pipeline.contracts.supply_monthly import (
    BLOCK_DEDUPLICATION_UNVERIFIED,
    SOURCE_REQUIRED_COLUMNS,
)
from data_pipeline.contracts.product_key import normalize_integer_code


ADAPTER_CONTRACT_VERSION: Final = "1.0.0"
DEFAULT_BATCH_SIZE: Final = 10_000
DEFAULT_HEADER_SCAN_LIMIT: Final = 12
DIAGNOSTIC_SAMPLE_LIMIT: Final = 20
SOURCE_BATCH_COLUMNS: Final[tuple[str, ...]] = (
    *SOURCE_REQUIRED_COLUMNS,
    "row_quality_flags",
)

DISCOVERY_REQUIRED_HEADERS: Final[frozenset[str]] = frozenset(
    {
        "공급일자",
        "공급구분",
        "공급한자 업체일련번호",
        "의료기기품목일련번호",
        "공급내역기준연월",
        "공급내역작업일련번호",
        "공급내역일련번호",
        "거래처 코드",
    }
)

HEADER_ALIASES: Final[dict[str, tuple[str, ...]]] = {
    "supply_date": ("공급일자",),
    "src_company_id": ("공급한자 업체일련번호",),
    "dst_company_id": ("공급받은자 업체일련번호",),
    "hospital_id": ("요양기관기호(의료기관)",),
    "item_serial": ("의료기기품목일련번호",),
    "model_serial": ("모델일련번호",),
    "udi_serial": ("UDI-DI 일련번호", "UDIDI일련번호"),
    "item_group_id": ("품목군",),
    "item_name_id": ("품목명",),
    "transaction_type": ("공급구분",),
    "amount_clean": ("공급금액",),
    "raw_supply_qty": ("공급수량",),
    "package_qty": ("포장내 총 수량",),
    "piece_qty": ("낱개총수량",),
    "udi": ("UDI-DI",),
    "supplier_type": ("업종",),
    "receiver_type": ("공급받은자업종",),
    "supplier_region": (
        "공급한자의 소재지 시도코드",
        "공급한자 시도코드",
    ),
    "receiver_region": (
        "공급받은자의 소재지 시도코드",
        "공급받은자 시도코드",
    ),
    "client_code": ("거래처 코드",),
    "base_month": ("공급내역기준연월",),
    "work_serial": ("공급내역작업일련번호",),
    "supply_serial": ("공급내역일련번호",),
    "reported_composite_key": ("공급내역보고자료복합Key",),
}
DISPLAY_NAME_ALIASES: Final[dict[str, tuple[str, ...]]] = {
    "supplier_display_name": ("공급자",),
    "receiver_display_name": ("공급받은자",),
}
CONSUMED_HEADERS: Final[frozenset[str]] = frozenset(
    alias
    for aliases in (*HEADER_ALIASES.values(), *DISPLAY_NAME_ALIASES.values())
    for alias in aliases
)
PROFILE_EXPECTED_FIELDS: Final[frozenset[str]] = frozenset(HEADER_ALIASES)
STRUCTURE_REQUIRED_FIELDS: Final[frozenset[str]] = frozenset(
    {
        "supply_date",
        "src_company_id",
        "transaction_type",
        "item_serial",
        "model_serial",
        "udi_serial",
        "raw_supply_qty",
        "client_code",
        "base_month",
        "work_serial",
        "supply_serial",
    }
)
RECEIVER_STRUCTURE_FIELDS: Final[frozenset[str]] = frozenset(
    {"dst_company_id", "hospital_id"}
)
TRANSACTION_TYPE_MAP: Final[dict[str, str]] = {
    "출고": "SUPPLY",
    "반품": "RETURN",
    "회수": "RECALL",
    "폐기": "DISCARD",
    "임대": "LEASE",
}
HIGH_VALUE_THRESHOLD: Final = Decimal("50000000")
BARCODE_SUSPECT_THRESHOLD: Final = Decimal("1e12")
_DECIMAL_PATTERN: Final = re.compile(
    r"^[+-]?(?:(?:\d+(?:\.\d*)?)|(?:\.\d+))(?:[eE][+-]?\d+)?$"
)
SUPPLY_WORKBOOK_RANGE_PATTERN: Final = re.compile(
    r"^공급내역보고자료\((\d{8})~(\d{8})\)\.xlsx$"
)


class NidsSupplyExcelError(RuntimeError):
    """Base error for the NIDS supply Excel adapter."""


class DataSheetDiscoveryError(NidsSupplyExcelError):
    """Raised when content-based sheet/header discovery is not unambiguous."""


class DataSheetSchemaError(DataSheetDiscoveryError):
    """Raised before row streaming when required logical fields are absent."""


class SourceSnapshotError(NidsSupplyExcelError):
    """Raised when source workbooks cannot form a deterministic snapshot."""


class SupplyWorkbookNameError(NidsSupplyExcelError):
    """Raised when a supply workbook name is not a closed dekade filename."""


@dataclass(frozen=True)
class SupplyWorkbookDateRange:
    month: str
    start: date
    end: date


@dataclass(frozen=True)
class ClosedSupplyMonth:
    month: str
    paths: tuple[Path, Path, Path]


@dataclass(frozen=True)
class RejectedSupplyMonth:
    month: str
    path_count: int
    logical_names: tuple[str, ...]
    reason: str


@dataclass(frozen=True)
class SupplyMonthGrouping:
    closed: tuple[ClosedSupplyMonth, ...]
    rejected: tuple[RejectedSupplyMonth, ...]


def _parse_yyyymmdd(value: str, *, field: str) -> date:
    """Parse a calendar day with datetime so leap years are not hard-coded."""
    try:
        parsed = datetime.strptime(value, "%Y%m%d").date()
    except ValueError as exc:
        raise SupplyWorkbookNameError(f"{field} is not a valid calendar day") from exc
    if parsed.strftime("%Y%m%d") != value:
        raise SupplyWorkbookNameError(f"{field} is not a valid calendar day")
    return parsed


def parse_supply_workbook_date_range(logical_name: str) -> SupplyWorkbookDateRange:
    """Read month membership from 공급내역보고자료(YYYYMMDD~YYYYMMDD).xlsx."""
    if not isinstance(logical_name, str) or not logical_name:
        raise SupplyWorkbookNameError("supply workbook logical name is required")
    match = SUPPLY_WORKBOOK_RANGE_PATTERN.fullmatch(logical_name)
    if match is None:
        raise SupplyWorkbookNameError(
            "supply workbook name must match 공급내역보고자료(YYYYMMDD~YYYYMMDD).xlsx"
        )
    start = _parse_yyyymmdd(match.group(1), field="range start")
    end = _parse_yyyymmdd(match.group(2), field="range end")
    if start.year != end.year or start.month != end.month:
        raise SupplyWorkbookNameError("supply workbook range must stay in one calendar month")
    if start > end:
        raise SupplyWorkbookNameError("supply workbook range start must not follow its end")
    last_day = date(start.year, start.month, monthrange(start.year, start.month)[1])
    if end > last_day:
        raise SupplyWorkbookNameError("supply workbook range end is past the calendar month")
    return SupplyWorkbookDateRange(start.strftime("%Y%m"), start, end)


def declared_month_from_logical_names(logical_names: Sequence[str]) -> str:
    """Return the single calendar month declared by dekade filenames."""
    if not logical_names:
        raise SupplyWorkbookNameError("at least one supply workbook name is required")
    months = tuple(parse_supply_workbook_date_range(name).month for name in logical_names)
    unique = tuple(sorted(set(months)))
    if len(unique) != 1:
        raise SupplyWorkbookNameError("supply workbooks must declare exactly one calendar month")
    return unique[0]


def group_closed_supply_months(paths: Sequence[Path]) -> SupplyMonthGrouping:
    """Group dekade files by month; exactly three files close a month."""
    if not paths:
        raise SupplyWorkbookNameError("at least one supply workbook path is required")
    by_month: dict[str, list[Path]] = defaultdict(list)
    for path in paths:
        if not isinstance(path, Path):
            raise TypeError("supply workbook paths must be pathlib.Path values")
        month = parse_supply_workbook_date_range(path.name).month
        by_month[month].append(path)
    closed: list[ClosedSupplyMonth] = []
    rejected: list[RejectedSupplyMonth] = []
    for month in sorted(by_month):
        month_paths = tuple(sorted(by_month[month], key=lambda item: item.name))
        names = tuple(path.name for path in month_paths)
        if len(month_paths) == 3:
            closed.append(ClosedSupplyMonth(month, (month_paths[0], month_paths[1], month_paths[2])))
            continue
        rejected.append(
            RejectedSupplyMonth(
                month=month,
                path_count=len(month_paths),
                logical_names=names,
                reason="expected_exactly_three_files",
            )
        )
    return SupplyMonthGrouping(tuple(closed), tuple(rejected))


@dataclass(frozen=True)
class WorkbookSnapshot:
    logical_name: str
    byte_size: int
    sha256: str


@dataclass(frozen=True)
class SourceLineage:
    adapter_contract_version: str
    source_version: str
    workbooks: tuple[WorkbookSnapshot, ...]

    def canonical_payload(self) -> dict[str, Any]:
        return {
            "adapter_contract_version": self.adapter_contract_version,
            "workbooks": [
                {
                    "byte_size": workbook.byte_size,
                    "logical_name": workbook.logical_name,
                    "sha256": workbook.sha256,
                }
                for workbook in self.workbooks
            ],
        }


@dataclass(frozen=True)
class DiscoveredSheet:
    name: str
    header_row: int
    headers: tuple[str, ...]


@dataclass
class IngestionIssue:
    status: str = ""
    total: int = 0
    sample: list[str] = field(default_factory=list)

    @property
    def omitted(self) -> int:
        return max(self.total - len(self.sample), 0)

    def add(self, location: str) -> None:
        self.total += 1
        if len(self.sample) < DIAGNOSTIC_SAMPLE_LIMIT:
            self.sample.append(location)


@dataclass
class SheetIngestionProfile:
    workbook: str
    sheet: str
    rows_read: int = 0
    rows_emitted: int = 0
    missing_columns: tuple[str, ...] = ()
    extra_columns: tuple[str, ...] = ()


@dataclass
class SupplyIngestionReport:
    sheet_profiles: list[SheetIngestionProfile] = field(default_factory=list)
    transaction_type_counts: Counter[str] = field(default_factory=Counter)
    source_identity_incomplete: IngestionIssue = field(
        default_factory=lambda: IngestionIssue(BLOCK_DEDUPLICATION_UNVERIFIED)
    )
    party_identity_incomplete: IngestionIssue = field(default_factory=IngestionIssue)
    product_key_incomplete: IngestionIssue = field(default_factory=IngestionIssue)
    date_conversion_failed: IngestionIssue = field(default_factory=IngestionIssue)
    amount_missing: IngestionIssue = field(default_factory=IngestionIssue)
    amount_conversion_failed: IngestionIssue = field(default_factory=IngestionIssue)
    quantity_missing: IngestionIssue = field(default_factory=IngestionIssue)
    quantity_conversion_failed: IngestionIssue = field(default_factory=IngestionIssue)
    piece_quantity_mismatch: IngestionIssue = field(default_factory=IngestionIssue)
    piece_quantity_unverified: IngestionIssue = field(default_factory=IngestionIssue)
    transaction_type_unknown: IngestionIssue = field(default_factory=IngestionIssue)
    reported_composite_key_available: int = 0
    reported_composite_key_missing: IngestionIssue = field(default_factory=IngestionIssue)
    reported_composite_key_present_identity_incomplete: IngestionIssue = field(
        default_factory=IngestionIssue
    )
    high_value_review: IngestionIssue = field(default_factory=IngestionIssue)
    barcode_entry_error_suspected: IngestionIssue = field(default_factory=IngestionIssue)
    high_value_max: Decimal | None = None
    rejected_by_reason: Counter[str] = field(default_factory=Counter)

    @property
    def rows_read(self) -> int:
        return sum(profile.rows_read for profile in self.sheet_profiles)

    @property
    def rows_emitted(self) -> int:
        return sum(profile.rows_emitted for profile in self.sheet_profiles)

    @property
    def rows_rejected(self) -> int:
        return sum(self.rejected_by_reason.values())

    @property
    def accounting_is_complete(self) -> bool:
        return (
            self.rows_read == self.rows_emitted + self.rows_rejected
            and self.rows_rejected == sum(self.rejected_by_reason.values())
        )

    def validate_accounting(self) -> None:
        if not self.accounting_is_complete:
            raise NidsSupplyExcelError(
                "Ingestion accounting invariant failed: rows_read must equal "
                "rows_emitted plus exclusive rows_rejected"
            )


def _canonical_json_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def _file_sha256(path: Path) -> str:
    digest = sha256()
    try:
        with path.open("rb") as stream:
            for block in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(block)
    except OSError as exc:
        raise SourceSnapshotError(
            f"Could not checksum source workbook {path.name!r}"
        ) from exc
    return digest.hexdigest()


def create_source_lineage(workbook_paths: Sequence[Path]) -> SourceLineage:
    """Hash a workbook set without including absolute paths or input order."""
    paths = tuple(workbook_paths)
    if not paths:
        raise SourceSnapshotError("At least one source workbook is required")
    if any(not isinstance(path, Path) for path in paths):
        raise TypeError("workbook_paths must contain pathlib.Path values")
    names = [path.name for path in paths]
    if len(set(names)) != len(names):
        raise SourceSnapshotError("Source workbook logical names must be unique")

    snapshots: list[WorkbookSnapshot] = []
    for path in paths:
        try:
            size = path.stat().st_size
        except OSError as exc:
            raise SourceSnapshotError(
                f"Could not stat source workbook {path.name!r}"
            ) from exc
        snapshots.append(WorkbookSnapshot(path.name, size, _file_sha256(path)))
    ordered = tuple(sorted(snapshots, key=lambda item: item.logical_name))
    payload = {
        "adapter_contract_version": ADAPTER_CONTRACT_VERSION,
        "workbooks": [
            {
                "byte_size": workbook.byte_size,
                "logical_name": workbook.logical_name,
                "sha256": workbook.sha256,
            }
            for workbook in ordered
        ],
    }
    source_version = f"nids-supply-v1:{sha256(_canonical_json_bytes(payload)).hexdigest()}"
    return SourceLineage(ADAPTER_CONTRACT_VERSION, source_version, ordered)


def _header_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _ignore_stored_worksheet_size(sheet: Any) -> None:
    """Ignore a stale xlsx dimension so read_only mode can see every used cell.

    NIDS exports often store ``A1`` as the worksheet size. openpyxl would then
    stop at column A and row 1 even when later cells exist in the XML.
    """
    reset = getattr(sheet, "reset_dimensions", None)
    if callable(reset):
        reset()


def _discover_open_workbook(
    workbook: Workbook,
    *,
    header_scan_limit: int,
) -> tuple[DiscoveredSheet, ...]:
    if header_scan_limit < 1 or header_scan_limit > 100:
        raise ValueError("header_scan_limit must be between 1 and 100")
    discovered: list[DiscoveredSheet] = []
    for sheet_name in sorted(workbook.sheetnames):
        sheet = workbook[sheet_name]
        _ignore_stored_worksheet_size(sheet)
        candidates: list[DiscoveredSheet] = []
        for row_number, row in enumerate(
            islice(sheet.iter_rows(values_only=True), header_scan_limit), start=1
        ):
            headers = tuple(_header_text(value) for value in row)
            present = {header for header in headers if header is not None}
            if not DISCOVERY_REQUIRED_HEADERS.issubset(present):
                continue
            nonempty = [header for header in headers if header is not None]
            if len(nonempty) != len(set(nonempty)):
                raise DataSheetDiscoveryError(
                    f"Duplicate headers in workbook sheet {sheet_name!r} at row {row_number}"
                )
            candidates.append(
                DiscoveredSheet(
                    name=sheet_name,
                    header_row=row_number,
                    headers=tuple(header or "" for header in headers),
                )
            )
        if len(candidates) > 1:
            rows = [candidate.header_row for candidate in candidates]
            raise DataSheetDiscoveryError(
                f"Ambiguous header rows in workbook sheet {sheet_name!r}: {rows}"
            )
        if candidates:
            discovered.append(candidates[0])
    if not discovered:
        raise DataSheetDiscoveryError("No data sheet contains the required header set")
    return tuple(discovered)


def discover_supply_sheets(
    workbook_path: Path,
    *,
    header_scan_limit: int = DEFAULT_HEADER_SCAN_LIMIT,
) -> tuple[DiscoveredSheet, ...]:
    """Discover all data sheets by bounded content inspection."""
    if not isinstance(workbook_path, Path):
        raise TypeError("workbook_path must be a pathlib.Path")
    workbook: Workbook | None = None
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        return _discover_open_workbook(
            workbook, header_scan_limit=header_scan_limit
        )
    except (DataSheetDiscoveryError, ValueError):
        raise
    except Exception as exc:
        raise NidsSupplyExcelError(
            f"Could not inspect source workbook {workbook_path.name!r}"
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _normalize_text(value: Any) -> str | None:
    if _is_missing(value):
        return None
    return str(value).strip()


def _normalize_display_name(value: Any) -> str | None:
    if _is_missing(value):
        return None
    collapsed = " ".join(str(value).split())
    if not collapsed:
        return None
    return collapsed[:200].rstrip() if len(collapsed) > 200 else collapsed


def _observe_display_name(
    counts: dict[str, Counter[str]],
    entity_id: str,
    raw_name: Any,
) -> None:
    name = _normalize_display_name(raw_name)
    if name is None:
        return
    counts.setdefault(entity_id, Counter())[name] += 1


def finalize_display_name_rows(
    counts: dict[str, Counter[str]],
) -> tuple[dict[str, Any], ...]:
    """Most-frequent display name per license ID; names are never identifiers."""
    rows: list[dict[str, Any]] = []
    for entity_id, name_counts in counts.items():
        ranked = sorted(name_counts.items(), key=lambda item: (-item[1], item[0]))
        rows.append(
            {
                "entity_id": entity_id,
                "display_name": ranked[0][0],
                "observation_count": int(sum(name_counts.values())),
                "distinct_name_count": int(len(name_counts)),
                "name_conflict": len(name_counts) > 1,
            }
        )
    rows.sort(key=lambda row: str(row["entity_id"]))
    return tuple(rows)


def _parse_decimal(value: Any) -> tuple[Decimal | None, str | None]:
    if _is_missing(value):
        return None, "missing"
    if isinstance(value, bool) or isinstance(value, float):
        return None, "invalid"
    if isinstance(value, Decimal):
        parsed = value
    elif isinstance(value, int):
        parsed = Decimal(value)
    else:
        text = str(value).strip()
        if not _DECIMAL_PATTERN.fullmatch(text):
            return None, "invalid"
        try:
            parsed = Decimal(text)
        except InvalidOperation:
            return None, "invalid"
    if not parsed.is_finite():
        return None, "invalid"
    if parsed < 0:
        return None, "negative"
    return parsed, None


def _parse_supply_date(value: Any) -> pd.Timestamp | None:
    if isinstance(value, (datetime, date, pd.Timestamp)):
        parsed = pd.Timestamp(value)
    else:
        text = _normalize_text(value)
        if text is None:
            return None
        try:
            parsed = pd.to_datetime(text, format="%Y%m%d", errors="raise")
        except (TypeError, ValueError):
            return None
    return parsed.normalize()


def _source_row_id(raw: dict[str, Any]) -> str | None:
    components = {
        "client_code": normalize_integer_code(raw.get("client_code")),
        "base_month": normalize_integer_code(raw.get("base_month")),
        "work_serial": normalize_integer_code(raw.get("work_serial")),
        "supply_serial": normalize_integer_code(raw.get("supply_serial")),
    }
    if any(value is None for value in components.values()):
        return None
    return f"nids-row-v1:{sha256(_canonical_json_bytes(components)).hexdigest()}"


def _alias_positions(
    headers: tuple[str, ...], aliases: dict[str, tuple[str, ...]]
) -> dict[str, int]:
    positions: dict[str, int] = {}
    for field_name, field_aliases in aliases.items():
        found = [headers.index(alias) for alias in field_aliases if alias in headers]
        if len(found) > 1:
            raise DataSheetDiscoveryError(
                f"Multiple aliases found for mapped field {field_name!r}"
            )
        if found:
            positions[field_name] = found[0]
    return positions


def _field_positions(headers: tuple[str, ...]) -> dict[str, int]:
    return {
        **_alias_positions(headers, HEADER_ALIASES),
        **_alias_positions(headers, DISPLAY_NAME_ALIASES),
    }


def _validate_mapped_sheet_schema(
    positions: dict[str, int],
    *,
    workbook_name: str,
    sheet_name: str,
) -> None:
    missing = set(STRUCTURE_REQUIRED_FIELDS - set(positions))
    if not RECEIVER_STRUCTURE_FIELDS.intersection(positions):
        missing.add("dst_company_id|hospital_id")
    if missing:
        raise DataSheetSchemaError(
            f"Workbook {workbook_name!r}, sheet {sheet_name!r} is missing "
            f"required logical fields: {sorted(missing)}"
        )


def _raw_fields(row: tuple[Any, ...], positions: dict[str, int]) -> dict[str, Any]:
    return {
        name: row[position] if position < len(row) else None
        for name, position in positions.items()
    }


def _location(workbook: str, sheet: str, row_number: int) -> str:
    return f"{workbook}:{sheet}:row={row_number}"


def _reject_row(
    report: SupplyIngestionReport,
    *,
    reason: str,
    issue: IngestionIssue,
    diagnostic: str,
) -> None:
    report.rejected_by_reason[reason] += 1
    issue.add(diagnostic)


def _map_row(
    raw: dict[str, Any],
    *,
    source_version: str,
    location: str,
    report: SupplyIngestionReport,
    display_name_counts: dict[str, Counter[str]],
) -> dict[str, Any] | None:
    source_row_id = _source_row_id(raw)
    diagnostic = source_row_id or location

    transaction_raw = _normalize_text(raw.get("transaction_type"))
    transaction_type = TRANSACTION_TYPE_MAP.get(transaction_raw or "")
    transaction_unknown = transaction_type is None
    if transaction_unknown:
        transaction_type = transaction_raw or "UNKNOWN"
        report.transaction_type_unknown.add(diagnostic)
    report.transaction_type_counts[
        "UNKNOWN" if transaction_unknown else transaction_type
    ] += 1

    reported_composite_present = (
        _normalize_text(raw.get("reported_composite_key")) is not None
    )
    if not reported_composite_present:
        report.reported_composite_key_missing.add(diagnostic)
    else:
        report.reported_composite_key_available += 1
        if source_row_id is None:
            report.reported_composite_key_present_identity_incomplete.add(location)

    amount, amount_error = _parse_decimal(raw.get("amount_clean"))
    raw_qty, raw_qty_error = _parse_decimal(raw.get("raw_supply_qty"))
    package_qty, package_qty_error = _parse_decimal(raw.get("package_qty"))
    official_piece, piece_error = _parse_decimal(raw.get("piece_qty"))
    if amount_error == "missing":
        report.amount_missing.add(diagnostic)
    elif amount_error is not None:
        report.amount_conversion_failed.add(diagnostic)
    for error in (raw_qty_error, package_qty_error, piece_error):
        if error == "missing":
            report.quantity_missing.add(diagnostic)
        elif error is not None:
            report.quantity_conversion_failed.add(diagnostic)

    if amount is not None and amount > HIGH_VALUE_THRESHOLD:
        report.high_value_review.add(diagnostic)
        report.high_value_max = (
            amount
            if report.high_value_max is None
            else max(report.high_value_max, amount)
        )
    if amount is not None and amount > BARCODE_SUSPECT_THRESHOLD:
        report.barcode_entry_error_suspected.add(diagnostic)

    if source_row_id is None:
        _reject_row(
            report,
            reason="source_identity_incomplete",
            issue=report.source_identity_incomplete,
            diagnostic=location,
        )
        return None

    src = normalize_integer_code(raw.get("src_company_id"))
    dst_company = normalize_integer_code(raw.get("dst_company_id"))
    hospital = _normalize_text(raw.get("hospital_id"))
    if src is None or (dst_company is None and hospital is None):
        _reject_row(
            report,
            reason="party_identity_incomplete",
            issue=report.party_identity_incomplete,
            diagnostic=source_row_id,
        )
        return None

    item_serial = normalize_integer_code(raw.get("item_serial"))
    model_serial = normalize_integer_code(raw.get("model_serial"))
    udi_serial = normalize_integer_code(raw.get("udi_serial"))
    if None in (item_serial, model_serial, udi_serial):
        _reject_row(
            report,
            reason="product_key_incomplete",
            issue=report.product_key_incomplete,
            diagnostic=source_row_id,
        )
        return None

    supply_date = _parse_supply_date(raw.get("supply_date"))
    if supply_date is None:
        _reject_row(
            report,
            reason="date_invalid",
            issue=report.date_conversion_failed,
            diagnostic=source_row_id,
        )
        return None

    flags: set[str] = set()
    if dst_company is None:
        dst_company_id = f"hosp:{hospital}"
        flags.add("receiver_hospital_fallback")
    else:
        dst_company_id = f"co:{dst_company}"

    if transaction_unknown:
        flags.add("transaction_type_unknown")

    if amount_error is not None:
        flags.add(f"amount_{amount_error}")
    if raw_qty_error is not None:
        flags.add(f"raw_supply_qty_{raw_qty_error}")
    if package_qty_error is not None:
        flags.add(f"package_qty_{package_qty_error}")
    if piece_error is not None:
        flags.add(f"piece_qty_{piece_error}")

    piece_qty: Decimal | None = None
    if official_piece is None or raw_qty is None or package_qty is None:
        report.piece_quantity_unverified.add(source_row_id)
        flags.add("piece_qty_unverified")
    elif official_piece != raw_qty * package_qty:
        report.piece_quantity_mismatch.add(source_row_id)
        flags.add("piece_qty_mismatch")
    else:
        piece_qty = official_piece

    if amount is not None and amount > HIGH_VALUE_THRESHOLD:
        flags.add("amount_high_value_review")
    if amount is not None and amount > BARCODE_SUSPECT_THRESHOLD:
        flags.add("amount_barcode_entry_error_suspected")

    src_company_id = f"co:{src}"
    _observe_display_name(
        display_name_counts, src_company_id, raw.get("supplier_display_name")
    )
    _observe_display_name(
        display_name_counts, dst_company_id, raw.get("receiver_display_name")
    )

    return {
        "supply_date": supply_date,
        "src_company_id": src_company_id,
        "dst_company_id": dst_company_id,
        "item_serial": item_serial,
        "model_serial": model_serial,
        "udi_serial": udi_serial,
        "item_group_id": _normalize_text(raw.get("item_group_id")),
        "item_name_id": _normalize_text(raw.get("item_name_id")),
        "transaction_type": transaction_type,
        "amount_clean": amount,
        "raw_supply_qty": raw_qty,
        "piece_qty": piece_qty,
        "udi": _normalize_text(raw.get("udi")),
        "supplier_type": _normalize_text(raw.get("supplier_type")),
        "receiver_type": _normalize_text(raw.get("receiver_type")),
        "supplier_region": _normalize_text(raw.get("supplier_region")),
        "receiver_region": _normalize_text(raw.get("receiver_region")),
        "source_version": source_version,
        "source_row_id": source_row_id,
        "row_quality_flags": ";".join(sorted(flags)),
    }


class SupplyExcelStream(Iterable[pd.DataFrame]):
    """One-pass bounded stream with lineage and a mutable bounded report."""

    def __init__(
        self,
        workbook_paths: Sequence[Path],
        *,
        batch_size: int,
        header_scan_limit: int,
        max_rows_per_workbook: int | None,
        create_source_lineage_snapshot: bool,
    ) -> None:
        if batch_size < 1:
            raise ValueError("batch_size must be positive")
        if max_rows_per_workbook is not None and max_rows_per_workbook < 1:
            raise ValueError("max_rows_per_workbook must be positive when set")
        self._paths = tuple(workbook_paths)
        if not self._paths:
            raise SourceSnapshotError("At least one source workbook is required")
        if any(not isinstance(path, Path) for path in self._paths):
            raise TypeError("workbook_paths must contain pathlib.Path values")
        names = [path.name for path in self._paths]
        if len(set(names)) != len(names):
            raise SourceSnapshotError("Source workbook logical names must be unique")
        self.batch_size = batch_size
        self.header_scan_limit = header_scan_limit
        self.max_rows_per_workbook = max_rows_per_workbook
        self.lineage: SourceLineage | None
        if create_source_lineage_snapshot:
            self.lineage = create_source_lineage(self._paths)
            path_by_name = {path.name: path for path in self._paths}
            self._ordered_paths = tuple(
                path_by_name[item.logical_name] for item in self.lineage.workbooks
            )
            self._source_version = self.lineage.source_version
        else:
            self.lineage = None
            self._ordered_paths = tuple(
                sorted(self._paths, key=lambda path: (path.name.casefold(), path.name))
            )
            self._source_version = "nids-supply-benchmark-unhashed-v1"
        self.report = SupplyIngestionReport()
        self._display_name_counts: dict[str, Counter[str]] = {}
        self._started = False
        self._closed = False
        self._active_generator: Iterator[pd.DataFrame] | None = None

    def display_name_rows(self) -> tuple[dict[str, Any], ...]:
        return finalize_display_name_rows(self._display_name_counts)

    def __enter__(self) -> SupplyExcelStream:
        if self._closed:
            raise RuntimeError("SupplyExcelStream is already closed")
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: Any,
    ) -> None:
        del exc_type, exc_value, traceback
        self.close()

    def __iter__(self) -> Iterator[pd.DataFrame]:
        if self._started or self._closed:
            raise RuntimeError("SupplyExcelStream is one-pass; create a new stream to reread")
        self._started = True
        generator = self._iter_batches()
        self._active_generator = generator
        return generator

    def close(self) -> None:
        """Close any active generator and its workbook; safe to call repeatedly."""
        if self._closed:
            return
        self._closed = True
        generator = self._active_generator
        self._active_generator = None
        if generator is not None:
            generator.close()

    def _iter_batches(self) -> Iterator[pd.DataFrame]:
        batch: list[dict[str, Any]] = []
        completed = False
        try:
            for path in self._ordered_paths:
                workbook: Workbook | None = None
                try:
                    workbook = load_workbook(path, read_only=True, data_only=True)
                    sheets = _discover_open_workbook(
                        workbook, header_scan_limit=self.header_scan_limit
                    )
                    workbook_rows_read = 0
                    for discovered in sheets:
                        positions = _field_positions(discovered.headers)
                        _validate_mapped_sheet_schema(
                            positions,
                            workbook_name=path.name,
                            sheet_name=discovered.name,
                        )
                        mapped_headers = {
                            discovered.headers[position]
                            for position in positions.values()
                        }
                        missing_fields = tuple(
                            sorted(
                                HEADER_ALIASES[field_name][0]
                                for field_name in PROFILE_EXPECTED_FIELDS
                                - set(positions)
                            )
                        )
                        profile = SheetIngestionProfile(
                            workbook=path.name,
                            sheet=discovered.name,
                            missing_columns=missing_fields,
                            extra_columns=tuple(
                                sorted(
                                    header
                                    for header in discovered.headers
                                    if header and header not in mapped_headers
                                )
                            ),
                        )
                        self.report.sheet_profiles.append(profile)
                        sheet = workbook[discovered.name]
                        _ignore_stored_worksheet_size(sheet)
                        for row_number, row in enumerate(
                            sheet.iter_rows(
                                min_row=discovered.header_row + 1,
                                values_only=True,
                            ),
                            start=discovered.header_row + 1,
                        ):
                            if not any(value is not None for value in row):
                                continue
                            if (
                                self.max_rows_per_workbook is not None
                                and workbook_rows_read >= self.max_rows_per_workbook
                            ):
                                break
                            workbook_rows_read += 1
                            profile.rows_read += 1
                            raw = _raw_fields(row, positions)
                            mapped = _map_row(
                                raw,
                                source_version=self._source_version,
                                location=_location(
                                    path.name, discovered.name, row_number
                                ),
                                report=self.report,
                                display_name_counts=self._display_name_counts,
                            )
                            if mapped is None:
                                continue
                            profile.rows_emitted += 1
                            batch.append(mapped)
                            if len(batch) == self.batch_size:
                                yield pd.DataFrame(
                                    batch, columns=SOURCE_BATCH_COLUMNS
                                )
                                batch = []
                        if (
                            self.max_rows_per_workbook is not None
                            and workbook_rows_read >= self.max_rows_per_workbook
                        ):
                            break
                except (DataSheetDiscoveryError, NidsSupplyExcelError):
                    raise
                except Exception as exc:
                    raise NidsSupplyExcelError(
                        f"Could not stream source workbook {path.name!r}"
                    ) from exc
                finally:
                    if workbook is not None:
                        workbook.close()
            if batch:
                yield pd.DataFrame(batch, columns=SOURCE_BATCH_COLUMNS)
            completed = True
        finally:
            self._active_generator = None
            self._closed = True
            if completed:
                self.report.validate_accounting()


def stream_nids_supply_excel(
    workbook_paths: Sequence[Path],
    *,
    batch_size: int = DEFAULT_BATCH_SIZE,
    header_scan_limit: int = DEFAULT_HEADER_SCAN_LIMIT,
    max_rows_per_workbook: int | None = None,
    create_source_lineage_snapshot: bool = True,
) -> SupplyExcelStream:
    """Create a one-pass bounded stream of normalized PR-01 source batches.

    ``create_source_lineage_snapshot=False`` is reserved for explicitly
    non-publishing benchmark reads. It avoids full workbook checksums and
    exposes no immutable lineage suitable for pipeline execution.
    """
    return SupplyExcelStream(
        workbook_paths,
        batch_size=batch_size,
        header_scan_limit=header_scan_limit,
        max_rows_per_workbook=max_rows_per_workbook,
        create_source_lineage_snapshot=create_source_lineage_snapshot,
    )
