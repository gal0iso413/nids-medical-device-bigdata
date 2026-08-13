"""Deterministic SQLite lookup for the official NIDS master product key."""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from hashlib import sha256
from itertools import islice
import json
from pathlib import Path, PurePosixPath
import re
import secrets
import shutil
import sqlite3
from typing import Any, Final, Iterable, Iterator, Sequence
from urllib.parse import quote

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from data_pipeline.contracts.product_key import normalize_integer_code
from data_pipeline.ingest.nids_supply_excel import SOURCE_BATCH_COLUMNS


DATASET_NAME: Final = "master_product_lookup"
LOGICAL_SCHEMA_NAME: Final = "nids_master_product_key"
LOGICAL_SCHEMA_VERSION: Final = "1.0.0"
STORAGE_CONTRACT_VERSION: Final = "1.0.0"
ADAPTER_CONTRACT_VERSION: Final = "1.0.0"
DATABASE_FILENAME: Final = "master_keys.sqlite"
MANIFEST_FILENAME: Final = "_manifest.json"
DEFAULT_HEADER_SCAN_LIMIT: Final = 12
DEFAULT_BATCH_SIZE: Final = 10_000
DIAGNOSTIC_SAMPLE_LIMIT: Final = 20
PRODUCT_KEY_COLUMNS: Final[tuple[str, str, str]] = (
    "item_serial",
    "model_serial",
    "udi_serial",
)
MASTER_HEADER_ALIASES: Final[dict[str, tuple[str, ...]]] = {
    "item_serial": ("의료기기품목일련번호",),
    "model_serial": ("모델일련번호",),
    "udi_serial": ("UDIDI일련번호",),
}
SQLITE_TABLE_SQL: Final = """CREATE TABLE product_key(
    item_serial TEXT NOT NULL,
    model_serial TEXT NOT NULL,
    udi_serial TEXT NOT NULL,
    PRIMARY KEY(item_serial, model_serial, udi_serial)
) WITHOUT ROWID"""
_SOURCE_HASH_PATTERN: Final = re.compile(r"^[0-9a-f]{64}$")
_MANIFEST_KEYS: Final = frozenset(
    {
        "adapter_contract_version",
        "database_file_size",
        "database_sha256",
        "dataset_name",
        "duplicate_key_rows",
        "invalid_key_rows",
        "invalid_key_locations",
        "invalid_key_omitted",
        "logical_schema_name",
        "logical_schema_version",
        "relative_database_path",
        "rows_read",
        "source_hash",
        "source_version",
        "source_workbooks",
        "storage_contract_version",
        "unique_key_count",
        "valid_key_rows",
    }
)


class MasterProductLookupError(RuntimeError):
    """Base error for master lookup creation, verification, and joins."""


class MasterSheetDiscoveryError(MasterProductLookupError):
    """Raised when master sheet discovery is missing or ambiguous."""


class MasterSheetSchemaError(MasterSheetDiscoveryError):
    """Raised when a likely master header lacks an official key field."""


class MasterSourceSnapshotError(MasterProductLookupError):
    """Raised when deterministic master lineage cannot be formed."""


class MasterLookupStorageError(MasterProductLookupError):
    """Raised for filesystem or SQLite I/O failures."""


class EmptyMasterLookupError(MasterProductLookupError):
    """Raised when a discovered master source contains no valid product key."""


class MasterLookupConflictError(MasterProductLookupError):
    """Raised instead of overwriting a different published artifact."""


class MasterLookupIntegrityError(MasterProductLookupError):
    """Raised when a published lookup is incomplete or corrupt."""


class SupplyBatchJoinError(MasterProductLookupError):
    """Raised when a supply batch cannot be joined under the exact contract."""


@dataclass(frozen=True)
class MasterWorkbookSnapshot:
    logical_name: str
    byte_size: int
    sha256: str


@dataclass(frozen=True)
class MasterSourceLineage:
    adapter_contract_version: str
    source_version: str
    source_hash: str
    workbooks: tuple[MasterWorkbookSnapshot, ...]

    def canonical_payload(self) -> dict[str, Any]:
        return {
            "adapter_contract_version": self.adapter_contract_version,
            "workbooks": [
                {
                    "byte_size": item.byte_size,
                    "logical_name": item.logical_name,
                    "sha256": item.sha256,
                }
                for item in self.workbooks
            ],
        }


@dataclass(frozen=True)
class MasterDiscoveredSheet:
    name: str
    header_row: int
    headers: tuple[str, ...]
    positions: tuple[int, int, int]


@dataclass
class MasterKeyIssue:
    total: int = 0
    sample: list[str] = field(default_factory=list)

    @property
    def omitted(self) -> int:
        return max(self.total - len(self.sample), 0)

    def add(self, location: str) -> None:
        self.total += 1
        if len(self.sample) < DIAGNOSTIC_SAMPLE_LIMIT:
            self.sample.append(location)


@dataclass
class MasterIngestionReport:
    rows_read: int = 0
    valid_key_rows: int = 0
    unique_key_count: int = 0
    invalid_key_rows: int = 0
    duplicate_key_rows: int = 0
    invalid_keys: MasterKeyIssue = field(default_factory=MasterKeyIssue)


@dataclass(frozen=True)
class MasterLookupBuildResult:
    status: str
    source_version: str
    source_hash: str
    relative_database_path: str
    rows_read: int
    valid_key_rows: int
    unique_key_count: int
    invalid_key_rows: int
    duplicate_key_rows: int
    invalid_key_locations: tuple[str, ...]
    invalid_key_omitted: int


@dataclass(frozen=True)
class MasterLookupVerification:
    source_version: str
    source_hash: str
    relative_database_path: str
    database_sha256: str
    database_file_size: int
    unique_key_count: int


@dataclass(frozen=True)
class MasterJoinReport:
    rows_input: int
    rows_matched: int
    rows_unmatched: int
    match_rate: Decimal | None
    unmatched_source_row_ids: tuple[str, ...]
    unmatched_omitted: int
    master_source_version: str
    supply_source_version: str


@dataclass(frozen=True)
class MasterJoinBatchResult:
    matched_rows: pd.DataFrame
    report: MasterJoinReport


def _canonical_json_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def _sha256_file(path: Path, *, operation: str) -> str:
    digest = sha256()
    try:
        with path.open("rb") as stream:
            for block in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(block)
    except OSError as exc:
        raise MasterLookupStorageError(f"Could not checksum file during {operation}") from exc
    return digest.hexdigest()


def create_master_lineage(workbook_paths: Sequence[Path]) -> MasterSourceLineage:
    """Create path- and input-order-independent master source lineage."""
    paths = tuple(workbook_paths)
    if not paths:
        raise MasterSourceSnapshotError("At least one master workbook is required")
    if any(not isinstance(path, Path) for path in paths):
        raise TypeError("workbook_paths must contain pathlib.Path values")
    if len({path.name for path in paths}) != len(paths):
        raise MasterSourceSnapshotError("Master workbook logical names must be unique")
    snapshots: list[MasterWorkbookSnapshot] = []
    for path in paths:
        try:
            size = path.stat().st_size
        except OSError as exc:
            raise MasterSourceSnapshotError(
                f"Could not stat master workbook {path.name!r}"
            ) from exc
        try:
            checksum = _sha256_file(path, operation="master lineage creation")
        except MasterProductLookupError as exc:
            raise MasterSourceSnapshotError(
                f"Could not checksum master workbook {path.name!r}"
            ) from exc
        snapshots.append(MasterWorkbookSnapshot(path.name, size, checksum))
    ordered = tuple(sorted(snapshots, key=lambda item: item.logical_name))
    payload = {
        "adapter_contract_version": ADAPTER_CONTRACT_VERSION,
        "workbooks": [
            {
                "byte_size": item.byte_size,
                "logical_name": item.logical_name,
                "sha256": item.sha256,
            }
            for item in ordered
        ],
    }
    source_hash = sha256(_canonical_json_bytes(payload)).hexdigest()
    return MasterSourceLineage(
        ADAPTER_CONTRACT_VERSION,
        f"nids-master-v1:{source_hash}",
        source_hash,
        ordered,
    )


def _header_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _field_positions(headers: tuple[str, ...], *, sheet_name: str) -> tuple[int, int, int]:
    result: list[int] = []
    for field in PRODUCT_KEY_COLUMNS:
        found = [headers.index(alias) for alias in MASTER_HEADER_ALIASES[field] if alias in headers]
        if len(found) > 1:
            raise MasterSheetDiscoveryError(
                f"Sheet {sheet_name!r} has multiple aliases for {field!r}"
            )
        if not found:
            raise MasterSheetSchemaError(
                f"Sheet {sheet_name!r} is missing required logical field {field!r}"
            )
        result.append(found[0])
    return tuple(result)  # type: ignore[return-value]


def _discover_open_master_workbook(
    workbook: Workbook,
    *,
    header_scan_limit: int,
) -> tuple[MasterDiscoveredSheet, ...]:
    if header_scan_limit < 1 or header_scan_limit > 100:
        raise ValueError("header_scan_limit must be between 1 and 100")
    discovered: list[MasterDiscoveredSheet] = []
    known_headers = {
        alias for aliases in MASTER_HEADER_ALIASES.values() for alias in aliases
    }
    for sheet_name in sorted(workbook.sheetnames):
        candidates: list[MasterDiscoveredSheet] = []
        for row_number, row in enumerate(
            islice(workbook[sheet_name].iter_rows(values_only=True), header_scan_limit),
            start=1,
        ):
            header_values = tuple(_header_text(value) for value in row)
            present = {value for value in header_values if value is not None}
            matched = present & known_headers
            if len(matched) < 2:
                continue
            nonempty = [value for value in header_values if value is not None]
            if len(nonempty) != len(set(nonempty)):
                raise MasterSheetDiscoveryError(
                    f"Duplicate headers in master sheet {sheet_name!r} at row {row_number}"
                )
            headers = tuple(value or "" for value in header_values)
            positions = _field_positions(headers, sheet_name=sheet_name)
            candidates.append(
                MasterDiscoveredSheet(sheet_name, row_number, headers, positions)
            )
        if len(candidates) > 1:
            raise MasterSheetDiscoveryError(
                f"Ambiguous header rows in master sheet {sheet_name!r}: "
                f"{[item.header_row for item in candidates]}"
            )
        if candidates:
            discovered.append(candidates[0])
    if not discovered:
        raise MasterSheetDiscoveryError(
            "No master data sheet contains the official three-key header set"
        )
    return tuple(discovered)


def discover_master_sheets(
    workbook_path: Path,
    *,
    header_scan_limit: int = DEFAULT_HEADER_SCAN_LIMIT,
) -> tuple[MasterDiscoveredSheet, ...]:
    """Discover every matching master sheet by bounded content inspection."""
    if not isinstance(workbook_path, Path):
        raise TypeError("workbook_path must be a pathlib.Path")
    workbook: Workbook | None = None
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        return _discover_open_master_workbook(
            workbook, header_scan_limit=header_scan_limit
        )
    except (MasterSheetDiscoveryError, ValueError):
        raise
    except Exception as exc:
        raise MasterProductLookupError(
            f"Could not inspect master workbook {workbook_path.name!r}"
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()


class MasterKeyStream(Iterable[tuple[str, str, str]]):
    """One-pass stream that closes active read-only workbooks promptly."""

    def __init__(
        self,
        workbook_paths: Sequence[Path],
        *,
        header_scan_limit: int = DEFAULT_HEADER_SCAN_LIMIT,
    ) -> None:
        self._paths = tuple(workbook_paths)
        self.lineage = create_master_lineage(self._paths)
        by_name = {path.name: path for path in self._paths}
        self._ordered_paths = tuple(by_name[item.logical_name] for item in self.lineage.workbooks)
        self.header_scan_limit = header_scan_limit
        self.report = MasterIngestionReport()
        self._started = False
        self._closed = False
        self._active_generator: Iterator[tuple[str, str, str]] | None = None

    def __enter__(self) -> MasterKeyStream:
        if self._closed:
            raise RuntimeError("MasterKeyStream is already closed")
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, traceback: Any) -> None:
        del exc_type, exc_value, traceback
        self.close()

    def __iter__(self) -> Iterator[tuple[str, str, str]]:
        if self._started or self._closed:
            raise RuntimeError("MasterKeyStream is one-pass; create a new stream to reread")
        self._started = True
        generator = self._iter_keys()
        self._active_generator = generator
        return generator

    def close(self) -> None:
        if self._closed:
            return
        self._closed = True
        generator = self._active_generator
        self._active_generator = None
        if generator is not None:
            generator.close()

    def _iter_keys(self) -> Iterator[tuple[str, str, str]]:
        try:
            for path in self._ordered_paths:
                workbook: Workbook | None = None
                try:
                    workbook = load_workbook(path, read_only=True, data_only=True)
                    sheets = _discover_open_master_workbook(
                        workbook, header_scan_limit=self.header_scan_limit
                    )
                    for discovered in sheets:
                        sheet = workbook[discovered.name]
                        for row_number, row in enumerate(
                            sheet.iter_rows(
                                min_row=discovered.header_row + 1,
                                values_only=True,
                            ),
                            start=discovered.header_row + 1,
                        ):
                            if not any(value is not None for value in row):
                                continue
                            self.report.rows_read += 1
                            values = tuple(
                                normalize_integer_code(
                                    row[position] if position < len(row) else None
                                )
                                for position in discovered.positions
                            )
                            if any(value is None for value in values):
                                self.report.invalid_key_rows += 1
                                self.report.invalid_keys.add(
                                    f"{path.name}:{discovered.name}:row={row_number}"
                                )
                                continue
                            self.report.valid_key_rows += 1
                            yield values  # type: ignore[misc]
                except (MasterSheetDiscoveryError, MasterProductLookupError):
                    raise
                except Exception as exc:
                    raise MasterProductLookupError(
                        f"Could not stream master workbook {path.name!r}"
                    ) from exc
                finally:
                    if workbook is not None:
                        workbook.close()
        finally:
            self._active_generator = None
            self._closed = True


def stream_master_product_keys(
    workbook_paths: Sequence[Path],
    *,
    header_scan_limit: int = DEFAULT_HEADER_SCAN_LIMIT,
) -> MasterKeyStream:
    return MasterKeyStream(workbook_paths, header_scan_limit=header_scan_limit)


def _lookup_dir(lookup_root: Path, source_hash: str) -> Path:
    if not isinstance(lookup_root, Path):
        raise TypeError("lookup_root must be a pathlib.Path")
    if not _SOURCE_HASH_PATTERN.fullmatch(source_hash):
        raise MasterProductLookupError("source_hash must be 64 lowercase hex characters")
    return (
        lookup_root
        / DATASET_NAME
        / f"schema_version={LOGICAL_SCHEMA_VERSION}"
        / f"source_hash={source_hash}"
    )


def _relative_database_path(source_hash: str) -> str:
    return str(
        PurePosixPath(
            DATASET_NAME,
            f"schema_version={LOGICAL_SCHEMA_VERSION}",
            f"source_hash={source_hash}",
            DATABASE_FILENAME,
        )
    )


def _manifest_from_report(
    lineage: MasterSourceLineage,
    report: MasterIngestionReport,
    database_path: Path,
) -> dict[str, Any]:
    try:
        size = database_path.stat().st_size
    except OSError as exc:
        raise MasterLookupStorageError("Could not stat candidate master lookup") from exc
    return {
        "adapter_contract_version": ADAPTER_CONTRACT_VERSION,
        "database_file_size": size,
        "database_sha256": _sha256_file(database_path, operation="manifest creation"),
        "dataset_name": DATASET_NAME,
        "duplicate_key_rows": report.duplicate_key_rows,
        "invalid_key_rows": report.invalid_key_rows,
        "invalid_key_locations": list(report.invalid_keys.sample),
        "invalid_key_omitted": report.invalid_keys.omitted,
        "logical_schema_name": LOGICAL_SCHEMA_NAME,
        "logical_schema_version": LOGICAL_SCHEMA_VERSION,
        "relative_database_path": _relative_database_path(lineage.source_hash),
        "rows_read": report.rows_read,
        "source_hash": lineage.source_hash,
        "source_version": lineage.source_version,
        "source_workbooks": [
            {
                "byte_size": item.byte_size,
                "logical_name": item.logical_name,
                "sha256": item.sha256,
            }
            for item in lineage.workbooks
        ],
        "storage_contract_version": STORAGE_CONTRACT_VERSION,
        "unique_key_count": report.unique_key_count,
        "valid_key_rows": report.valid_key_rows,
    }


def _result_from_manifest(status: str, manifest: dict[str, Any]) -> MasterLookupBuildResult:
    return MasterLookupBuildResult(
        status=status,
        source_version=manifest["source_version"],
        source_hash=manifest["source_hash"],
        relative_database_path=manifest["relative_database_path"],
        rows_read=manifest["rows_read"],
        valid_key_rows=manifest["valid_key_rows"],
        unique_key_count=manifest["unique_key_count"],
        invalid_key_rows=manifest["invalid_key_rows"],
        duplicate_key_rows=manifest["duplicate_key_rows"],
        invalid_key_locations=tuple(manifest["invalid_key_locations"]),
        invalid_key_omitted=manifest["invalid_key_omitted"],
    )


def _parse_manifest(final_dir: Path, source_hash: str) -> dict[str, Any]:
    path = final_dir / MANIFEST_FILENAME
    if not path.is_file():
        raise MasterLookupIntegrityError("Master lookup manifest is missing")
    try:
        raw = path.read_bytes()
        manifest = json.loads(raw.decode("utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise MasterLookupIntegrityError("Master lookup manifest is unreadable") from exc
    if not isinstance(manifest, dict) or frozenset(manifest) != _MANIFEST_KEYS:
        raise MasterLookupIntegrityError("Master lookup manifest schema is invalid")
    if raw != _canonical_json_bytes(manifest):
        raise MasterLookupIntegrityError("Master lookup manifest is not canonical JSON")
    expected = {
        "dataset_name": DATASET_NAME,
        "logical_schema_name": LOGICAL_SCHEMA_NAME,
        "logical_schema_version": LOGICAL_SCHEMA_VERSION,
        "storage_contract_version": STORAGE_CONTRACT_VERSION,
        "adapter_contract_version": ADAPTER_CONTRACT_VERSION,
        "source_hash": source_hash,
        "source_version": f"nids-master-v1:{source_hash}",
        "relative_database_path": _relative_database_path(source_hash),
    }
    for key, value in expected.items():
        if manifest.get(key) != value:
            raise MasterLookupIntegrityError(f"Master lookup manifest has invalid {key}")
    for key in (
        "rows_read",
        "valid_key_rows",
        "unique_key_count",
        "invalid_key_rows",
        "duplicate_key_rows",
        "database_file_size",
        "invalid_key_omitted",
    ):
        if not isinstance(manifest.get(key), int) or manifest[key] < 0:
            raise MasterLookupIntegrityError(f"Master lookup manifest has invalid {key}")
    locations = manifest.get("invalid_key_locations")
    if (
        not isinstance(locations, list)
        or len(locations) > DIAGNOSTIC_SAMPLE_LIMIT
        or any(not isinstance(value, str) or len(value) > 512 for value in locations)
        or manifest["invalid_key_omitted"]
        != max(manifest["invalid_key_rows"] - len(locations), 0)
    ):
        raise MasterLookupIntegrityError("Master lookup invalid-key diagnostics are invalid")
    if (
        not isinstance(manifest.get("database_sha256"), str)
        or not _SOURCE_HASH_PATTERN.fullmatch(manifest["database_sha256"])
    ):
        raise MasterLookupIntegrityError("Master lookup manifest has invalid database_sha256")
    if manifest["rows_read"] != manifest["valid_key_rows"] + manifest["invalid_key_rows"]:
        raise MasterLookupIntegrityError("Master lookup row accounting is invalid")
    if manifest["valid_key_rows"] != manifest["unique_key_count"] + manifest["duplicate_key_rows"]:
        raise MasterLookupIntegrityError("Master lookup duplicate accounting is invalid")
    workbooks = manifest.get("source_workbooks")
    if not isinstance(workbooks, list) or not workbooks:
        raise MasterLookupIntegrityError("Master lookup source_workbooks are invalid")
    logical_names: list[str] = []
    for workbook in workbooks:
        if not isinstance(workbook, dict) or frozenset(workbook) != {
            "byte_size", "logical_name", "sha256"
        }:
            raise MasterLookupIntegrityError("Master lookup source workbook entry is invalid")
        if (
            not isinstance(workbook["logical_name"], str)
            or not workbook["logical_name"]
            or Path(workbook["logical_name"]).name != workbook["logical_name"]
            or not isinstance(workbook["byte_size"], int)
            or workbook["byte_size"] < 0
            or not isinstance(workbook["sha256"], str)
            or not _SOURCE_HASH_PATTERN.fullmatch(workbook["sha256"])
        ):
            raise MasterLookupIntegrityError("Master lookup source workbook entry is invalid")
        logical_names.append(workbook["logical_name"])
    if logical_names != sorted(logical_names) or len(logical_names) != len(set(logical_names)):
        raise MasterLookupIntegrityError("Master lookup source workbooks are not canonical")
    lineage_payload = {
        "adapter_contract_version": manifest["adapter_contract_version"],
        "workbooks": workbooks,
    }
    if sha256(_canonical_json_bytes(lineage_payload)).hexdigest() != source_hash:
        raise MasterLookupIntegrityError("Master lookup source lineage hash is invalid")
    return manifest


def _connect_read_only(database_path: Path) -> sqlite3.Connection:
    uri_path = quote(database_path.resolve().as_posix(), safe="/:/")
    return sqlite3.connect(f"file:{uri_path}?mode=ro", uri=True)


def verify_master_product_lookup(
    lookup_root: Path,
    source_hash: str,
) -> MasterLookupVerification:
    """Verify canonical manifest, checksum, schema, and key count."""
    final_dir = _lookup_dir(lookup_root, source_hash)
    if not final_dir.is_dir():
        raise MasterLookupIntegrityError("Master lookup directory is missing")
    manifest = _parse_manifest(final_dir, source_hash)
    database_path = final_dir / DATABASE_FILENAME
    if not database_path.is_file():
        raise MasterLookupIntegrityError("Master lookup database is missing")
    try:
        size = database_path.stat().st_size
    except OSError as exc:
        raise MasterLookupStorageError("Could not stat master lookup database") from exc
    if size != manifest["database_file_size"]:
        raise MasterLookupIntegrityError("Master lookup database size does not match manifest")
    checksum = _sha256_file(database_path, operation="lookup verification")
    if checksum != manifest["database_sha256"]:
        raise MasterLookupIntegrityError("Master lookup database checksum does not match manifest")
    connection: sqlite3.Connection | None = None
    try:
        connection = _connect_read_only(database_path)
        integrity = connection.execute("PRAGMA quick_check").fetchone()
        if integrity != ("ok",):
            raise MasterLookupIntegrityError("Master lookup SQLite integrity check failed")
        table = connection.execute(
            "SELECT sql FROM sqlite_master WHERE type='table' AND name='product_key'"
        ).fetchone()
        if table is None or "WITHOUT ROWID" not in str(table[0]).upper():
            raise MasterLookupIntegrityError("Master lookup product_key schema is invalid")
        columns = connection.execute("PRAGMA table_info(product_key)").fetchall()
        if (
            [row[1] for row in columns] != list(PRODUCT_KEY_COLUMNS)
            or [str(row[2]).upper() for row in columns] != ["TEXT", "TEXT", "TEXT"]
            or [row[3] for row in columns] != [1, 1, 1]
            or [row[5] for row in columns] != [1, 2, 3]
        ):
            raise MasterLookupIntegrityError("Master lookup product_key columns are invalid")
        count = int(connection.execute("SELECT COUNT(*) FROM product_key").fetchone()[0])
    except sqlite3.Error as exc:
        raise MasterLookupIntegrityError("Master lookup SQLite content is unreadable") from exc
    finally:
        if connection is not None:
            connection.close()
    if count != manifest["unique_key_count"]:
        raise MasterLookupIntegrityError("Master lookup key count does not match manifest")
    return MasterLookupVerification(
        manifest["source_version"],
        source_hash,
        manifest["relative_database_path"],
        checksum,
        size,
        count,
    )


def _write_candidate_lookup(
    workbook_paths: Sequence[Path],
    lineage: MasterSourceLineage,
    temp_dir: Path,
    *,
    header_scan_limit: int,
    batch_size: int,
) -> dict[str, Any]:
    database_path = temp_dir / DATABASE_FILENAME
    stream = MasterKeyStream(workbook_paths, header_scan_limit=header_scan_limit)
    if stream.lineage != lineage:
        raise MasterSourceSnapshotError("Master source changed after lineage creation")
    connection: sqlite3.Connection | None = None
    try:
        connection = sqlite3.connect(database_path)
        connection.execute("PRAGMA journal_mode=OFF")
        connection.execute("PRAGMA synchronous=OFF")
        connection.execute("PRAGMA page_size=4096")
        connection.execute(SQLITE_TABLE_SQL)
        batch: list[tuple[str, str, str]] = []
        with stream:
            for key in stream:
                batch.append(key)
                if len(batch) == batch_size:
                    before = connection.total_changes
                    connection.executemany(
                        "INSERT OR IGNORE INTO product_key VALUES (?, ?, ?)", batch
                    )
                    stream.report.unique_key_count += connection.total_changes - before
                    batch.clear()
            if batch:
                before = connection.total_changes
                connection.executemany(
                    "INSERT OR IGNORE INTO product_key VALUES (?, ?, ?)", batch
                )
                stream.report.unique_key_count += connection.total_changes - before
        stream.report.duplicate_key_rows = (
            stream.report.valid_key_rows - stream.report.unique_key_count
        )
        if stream.report.unique_key_count == 0:
            raise EmptyMasterLookupError(
                "Master source contains no valid official three-key product identity"
            )
        connection.commit()
        connection.execute("VACUUM")
    except sqlite3.Error as exc:
        raise MasterLookupStorageError("Could not build master lookup SQLite database") from exc
    finally:
        if connection is not None:
            connection.close()
        stream.close()
    manifest = _manifest_from_report(lineage, stream.report, database_path)
    try:
        (temp_dir / MANIFEST_FILENAME).write_bytes(_canonical_json_bytes(manifest))
    except OSError as exc:
        raise MasterLookupStorageError("Could not write master lookup manifest") from exc
    return manifest


def _candidate_matches(
    candidate: dict[str, Any], verification: MasterLookupVerification
) -> bool:
    return (
        candidate["database_sha256"] == verification.database_sha256
        and candidate["database_file_size"] == verification.database_file_size
        and candidate["unique_key_count"] == verification.unique_key_count
    )


def _publish_candidate_lookup(
    temp_dir: Path,
    final_dir: Path,
    lookup_root: Path,
    source_hash: str,
    manifest: dict[str, Any],
) -> bool:
    try:
        temp_dir.replace(final_dir)
        return True
    except OSError as publish_exc:
        if not final_dir.is_dir():
            raise MasterLookupStorageError("Could not publish master lookup") from publish_exc
        raced = verify_master_product_lookup(lookup_root, source_hash)
        if _candidate_matches(manifest, raced):
            return False
        raise MasterLookupConflictError(
            "Master lookup appeared with different content during publication"
        ) from publish_exc


def build_master_product_lookup(
    workbook_paths: Sequence[Path],
    lookup_root: Path,
    *,
    header_scan_limit: int = DEFAULT_HEADER_SCAN_LIMIT,
    batch_size: int = DEFAULT_BATCH_SIZE,
) -> MasterLookupBuildResult:
    """Build and atomically publish one immutable source-version lookup."""
    if batch_size < 1:
        raise ValueError("batch_size must be positive")
    lineage = create_master_lineage(workbook_paths)
    final_dir = _lookup_dir(lookup_root, lineage.source_hash)
    if final_dir.exists():
        verify_master_product_lookup(lookup_root, lineage.source_hash)
        manifest = _parse_manifest(final_dir, lineage.source_hash)
        return _result_from_manifest("unchanged", manifest)
    parent = final_dir.parent
    staging_root = lookup_root / DATASET_NAME
    try:
        parent.mkdir(parents=True, exist_ok=True)
        staging_root.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise MasterLookupStorageError("Could not create master lookup storage root") from exc
    # Keep the private staging name short so Windows callers do not consume an
    # extra full source hash before the atomic move into the contracted path.
    temp_dir = staging_root / f".lookup.tmp-{secrets.token_hex(8)}"
    try:
        try:
            temp_dir.mkdir()
        except OSError as exc:
            raise MasterLookupStorageError(
                "Could not create master lookup staging directory"
            ) from exc
        manifest = _write_candidate_lookup(
            workbook_paths,
            lineage,
            temp_dir,
            header_scan_limit=header_scan_limit,
            batch_size=batch_size,
        )
        published = _publish_candidate_lookup(
            temp_dir, final_dir, lookup_root, lineage.source_hash, manifest
        )
        return _result_from_manifest("written" if published else "unchanged", manifest)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def _select_matching_batch_positions(
    connection: sqlite3.Connection,
    table: str,
) -> list[int]:
    return [
        int(row[0])
        for row in connection.execute(
            f"SELECT b.position FROM {table} b "
            "JOIN product_key p USING(item_serial, model_serial, udi_serial) "
            "ORDER BY b.position"
        )
    ]


class MasterProductLookup:
    """Verified read-only lookup with bounded temporary-table joins."""

    def __init__(self, lookup_root: Path, source_hash: str) -> None:
        verification = verify_master_product_lookup(lookup_root, source_hash)
        self.verification = verification
        database_path = _lookup_dir(lookup_root, source_hash) / DATABASE_FILENAME
        self._connection = _connect_read_only(database_path)
        self._closed = False

    def __enter__(self) -> MasterProductLookup:
        if self._closed:
            raise RuntimeError("MasterProductLookup is closed")
        return self

    def __exit__(self, exc_type: Any, exc_value: Any, traceback: Any) -> None:
        del exc_type, exc_value, traceback
        self.close()

    def close(self) -> None:
        if self._closed:
            return
        self._closed = True
        self._connection.close()

    def join_supply_batch(self, batch: pd.DataFrame) -> MasterJoinBatchResult:
        if self._closed:
            raise RuntimeError("MasterProductLookup is closed")
        if not isinstance(batch, pd.DataFrame):
            raise SupplyBatchJoinError("Supply batch must be a pandas DataFrame")
        if tuple(batch.columns) != SOURCE_BATCH_COLUMNS:
            raise SupplyBatchJoinError("Supply batch columns must exactly match SOURCE_BATCH_COLUMNS")
        if batch["source_row_id"].isna().any() or batch["source_row_id"].astype("string").str.strip().eq("").any():
            raise SupplyBatchJoinError("Supply batch contains an invalid source_row_id")
        source_versions = batch["source_version"].astype("string").str.strip()
        versions = source_versions.dropna().unique().tolist()
        if (
            len(versions) != 1
            or source_versions.isna().any()
            or source_versions.eq("").any()
        ):
            raise SupplyBatchJoinError("Supply batch must contain exactly one source_version")
        normalized_keys: list[tuple[str, str, str]] = []
        for values in batch.loc[:, PRODUCT_KEY_COLUMNS].itertuples(index=False, name=None):
            normalized = tuple(normalize_integer_code(value) for value in values)
            if any(value is None for value in normalized):
                raise SupplyBatchJoinError("Supply batch contains an incomplete official three-key")
            normalized_keys.append(normalized)  # type: ignore[arg-type]
        table = f"batch_product_key_{secrets.token_hex(8)}"
        try:
            self._connection.execute("BEGIN")
            self._connection.execute(
                f"CREATE TEMP TABLE {table}(position INTEGER PRIMARY KEY, item_serial TEXT NOT NULL, model_serial TEXT NOT NULL, udi_serial TEXT NOT NULL)"
            )
            self._connection.executemany(
                f"INSERT INTO {table} VALUES (?, ?, ?, ?)",
                (
                    (position, *key)
                    for position, key in enumerate(normalized_keys)
                ),
            )
            matched_positions = _select_matching_batch_positions(
                self._connection, table
            )
            self._connection.execute(f"DROP TABLE {table}")
            self._connection.commit()
        except sqlite3.Error as exc:
            cleanup_errors: list[sqlite3.Error] = []
            try:
                if self._connection.in_transaction:
                    self._connection.rollback()
            except sqlite3.Error as cleanup_exc:
                cleanup_errors.append(cleanup_exc)
            try:
                self._connection.execute(f"DROP TABLE IF EXISTS {table}")
                if self._connection.in_transaction:
                    self._connection.commit()
            except sqlite3.Error as cleanup_exc:
                cleanup_errors.append(cleanup_exc)
                try:
                    if self._connection.in_transaction:
                        self._connection.rollback()
                except sqlite3.Error as rollback_exc:
                    cleanup_errors.append(rollback_exc)
            error = SupplyBatchJoinError("Could not join supply batch to master lookup")
            for cleanup_error in cleanup_errors:
                error.add_note(
                    f"Temporary batch cleanup also failed: {cleanup_error.__class__.__name__}"
                )
            raise error from exc
        matched_set = set(matched_positions)
        unmatched_positions = (
            position for position in range(len(batch)) if position not in matched_set
        )
        sample_positions = list(islice(unmatched_positions, DIAGNOSTIC_SAMPLE_LIMIT))
        samples = tuple(str(batch.iloc[position]["source_row_id"]) for position in sample_positions)
        matched_count = len(matched_positions)
        unmatched_count = len(batch) - matched_count
        return MasterJoinBatchResult(
            batch.iloc[matched_positions].copy(deep=True),
            MasterJoinReport(
                rows_input=len(batch),
                rows_matched=matched_count,
                rows_unmatched=unmatched_count,
                match_rate=(
                    Decimal(matched_count) / Decimal(len(batch)) if len(batch) else None
                ),
                unmatched_source_row_ids=samples,
                unmatched_omitted=max(unmatched_count - len(samples), 0),
                master_source_version=self.verification.source_version,
                supply_source_version=versions[0],
            ),
        )


def open_master_product_lookup(
    lookup_root: Path, source_hash: str
) -> MasterProductLookup:
    return MasterProductLookup(lookup_root, source_hash)


def join_supply_batch_to_master(
    batch: pd.DataFrame,
    lookup: MasterProductLookup,
) -> MasterJoinBatchResult:
    """Join one bounded PR-03A batch through the verified read-only lookup."""
    if not isinstance(lookup, MasterProductLookup):
        raise TypeError("lookup must be a MasterProductLookup")
    return lookup.join_supply_batch(batch)
